"""Focused manual-input bot data extractor.

This is the canonical extractor entrypoint. Shared helper logic is imported
from a private compatibility core so the public runtime stays decoupled from
the deprecated wrapper module.
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import sys
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

import numpy as np  # pyright: ignore[reportMissingImports]
import pandas as pd  # pyright: ignore[reportMissingImports]

from neutralgrid.api.binance_client import BinanceClient
from neutralgrid.data.funding_rate import realized_funding_carry_proxy_next_hours
from neutralgrid.scanner.feature_extractor import compute_profile_feature_block

from neutralgrid.metrics.pnl_curve_features_v20260310 import extract_pnl_curve_features

from _bot_data_extractor_core import (
    ExtractedBotData,
    ParsedTradeFill,
    TechnicalIndicators,
    TradeMetrics,
    _compute_liquidation_features,
    auto_repair_coherence,
    build_bot_data_from_text,
    compute_indicators,
    compute_pnl_curve_excursions,
    compute_trade_metrics,
    extract_funding_fee_events_from_transactions,
    fetch_historical_klines,
    fetch_session_mark_klines,
    filter_trades_by_symbol_and_time,
    load_trade_history,
    load_transaction_history,
    parse_pnl_curve_from_text,
    parse_manual_ui_datetime_to_utc,
    parse_screenshot_text,
    parse_trade_fills_from_text,
    parse_user_text,
    prepare_entry_validation_row,
    search_candidate_match,
    text_fills_to_dataframe,
    validate_data_coherence,
)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Preserve legacy/public re-exports used by tests and compatibility scripts.
_COMPAT_REEXPORTS = (
    TechnicalIndicators,
    _compute_liquidation_features,
    parse_pnl_curve_from_text,
    parse_screenshot_text,
    parse_trade_fills_from_text,
    parse_user_text,
)


MANUAL_INPUT_ARCHIVE_DIR = Path("data/manual_input")


def _persist_manual_input_text(
    text: str,
    symbol: str,
    strategy_id: Any,
    *,
    base_dir: Path = MANUAL_INPUT_ARCHIVE_DIR,
) -> Optional[Path]:
    """Archive the raw manual input under data/manual_input/<YYYY-MM-DD>/<SYMBOL>_<strategy_id>.txt.

    Idempotent: if the target file already exists, the archive is left untouched.
    Returns the written path, the existing path on no-op, or None on failure.
    """
    if not symbol or strategy_id is None:
        return None
    try:
        date_dir = base_dir / datetime.now().strftime("%Y-%m-%d")
        date_dir.mkdir(parents=True, exist_ok=True)
        target = date_dir / f"{symbol}_{strategy_id}.txt"
        if target.exists():
            logger.info("Manual input already archived at %s; skipping rewrite", target)
            return target
        target.write_text(text, encoding="utf-8")
        logger.info("Archived manual input → %s", target)
        return target
    except OSError as exc:
        logger.warning("Failed to archive manual input for %s/%s: %s", symbol, strategy_id, exc)
        return None


DEFAULT_WORKBOOK_COLUMNS: List[str] = [
    "strategy_id",
    "symbol",
    "strategy_type",
    "status",
    "start_time_utc",
    "end_time_utc",
    "duration_hours",
    "invested_margin_usdt",
    "leverage",
    "grids_count",
    "price_range_low",
    "price_range_high",
    "grid_spacing_pct",
    "total_profit_usdt",
    "pnl_pct",
    "realized_pnl_usdt",
    "unrealized_pnl_usdt",
    "funding_fee_usdt",
    "commission_usdt",
    "profit_factor",
    "maker_count",
    "taker_count",
    "total_trades",
    "executed_qty",
    "notional_completed",
    "holding_time_ok",
    "last_updated_utc",
    "adx_1h",
    "adx_15m",
    "adx_5m",
    "rsi_15m",
    "ema_slope_1h",
    "ema_crosses_5m",
    "vwap_crosses_5m",
    "range_size_pct",
    "bb_width",
    "trend_structure",
    "candidate_id",
    "mae",
    "mfe",
    "mae_pct_initial",
    "mfe_pct_initial",
    "liq_price_long",
    "liq_price_short",
    "dist_to_liq_long_pct",
    "dist_to_liq_short_pct",
    "liq_range_utilization",
    "liq_asymmetry",
    "trigger_price",
    "parkinson_vol_ratio_4h_24h_pre",
    "variance_ratio_1m_15m_pre_2h",
    "funding_carry_expected_next_7h",
    "liquidity_stability_z_1h",
    "mode",
]


@dataclass(frozen=True)
class MatchedProfitEvent:
    """A realized matched-profit event from the Binance manual export."""

    timestamp: datetime
    profit_usdt: float


def parse_matched_profit_events(text: str) -> List[MatchedProfitEvent]:
    """Extract realized matched-profit entries from the manual text dump."""
    pattern = re.compile(
        r"(?m)^(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+([-\d,.]+)\s*USDT\s*$",
        re.IGNORECASE,
    )

    events: List[MatchedProfitEvent] = []
    for timestamp_str, profit_str in pattern.findall(text):
        events.append(
            MatchedProfitEvent(
                timestamp=parse_manual_ui_datetime_to_utc(timestamp_str),
                profit_usdt=float(profit_str.replace(",", "")),
            )
        )
    return events


def _require_active_worksheet(workbook: Any) -> Any:
    """Return the active worksheet with an explicit runtime guarantee."""
    worksheet = workbook.active
    if worksheet is None:
        raise ValueError("Workbook has no active worksheet.")
    return worksheet


def resolve_workbook_columns(output_path: Path, *, read_only: bool = False) -> List[str]:
    """Return the current workbook schema without mutating an existing file.

    Parameters
    ----------
    output_path
        Path to the workbook file.
    read_only
        Retained for API compatibility. Existing workbooks are never widened.
    """
    if not output_path.exists():
        return list(DEFAULT_WORKBOOK_COLUMNS)

    import openpyxl  # pyright: ignore[reportMissingModuleSource]

    workbook = openpyxl.load_workbook(output_path)
    try:
        worksheet = _require_active_worksheet(workbook)
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell) for cell in header_row if cell is not None]
    finally:
        workbook.close()

    header_set = set(headers)
    default_set = set(DEFAULT_WORKBOOK_COLUMNS)
    extras = [col for col in headers if col not in default_set]

    if extras:
        raise ValueError(
            f"Workbook schema drift detected. Extra columns not in DEFAULT_WORKBOOK_COLUMNS: {extras}"
        )

    missing_defaults = [col for col in DEFAULT_WORKBOOK_COLUMNS if col not in header_set]
    if missing_defaults:
        logger.warning(
            "Workbook is missing %d default column(s); preserving existing schema: %s",
            len(missing_defaults),
            missing_defaults,
        )

    return headers


def filter_row_to_schema(row: Dict[str, Any], columns: Sequence[str]) -> Dict[str, Any]:
    """Keep only workbook columns and preserve their order."""
    return {column: row.get(column) for column in columns}


def _expand_worksheet_tables_to_data(worksheet: Any) -> None:
    """Keep imported Excel table references aligned after an append.

    The canonical workbook uses named tables for filtering and visual review.
    ``openpyxl.Worksheet.append`` writes the row but does not expand those table
    references automatically, leaving the new row outside the canonical table.
    Only tables anchored at A1 and spanning the worksheet's active columns are
    extended; unrelated tables are left untouched.
    """
    from openpyxl.utils import get_column_letter, range_boundaries  # pyright: ignore[reportMissingModuleSource]

    for table in worksheet.tables.values():
        min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
        if min_row != 1 or min_col != 1 or max_col != worksheet.max_column:
            continue
        table.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"


def append_row_preserving_schema(output_path: Path, row: Dict[str, Any], columns: Sequence[str]) -> bool:
    """Append a row without changing workbook headers or order."""
    import openpyxl  # pyright: ignore[reportMissingModuleSource]

    workbook_exists = output_path.exists()
    workbook = (
        openpyxl.load_workbook(output_path)
        if workbook_exists
        else openpyxl.Workbook()
    )

    try:
        worksheet = _require_active_worksheet(workbook)
        if not workbook_exists:
            worksheet.title = "Sheet1"
            for column_index, column_name in enumerate(columns, start=1):
                worksheet.cell(row=1, column=column_index, value=column_name)

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell) for cell in header_row if cell is not None]
        if headers != list(columns):
            raise ValueError("Workbook header order does not match the active schema.")

        try:
            strategy_id_index = headers.index("strategy_id") + 1
        except ValueError as exc:
            raise ValueError("Workbook is missing strategy_id header.") from exc

        strategy_id = row.get("strategy_id")
        if strategy_id is not None:
            for existing in worksheet.iter_rows(
                min_row=2,
                min_col=strategy_id_index,
                max_col=strategy_id_index,
                values_only=True,
            ):
                if existing[0] == strategy_id:
                    logger.warning("strategy_id %s already exists; skipping append", strategy_id)
                    return False

        row_values: List[Any] = []
        for column_name in columns:
            value = row.get(column_name)
            if isinstance(value, datetime) and value.tzinfo is not None:
                value = value.replace(tzinfo=None)
            if isinstance(value, (bool, np.bool_)):
                value = int(value)
            row_values.append(value)

        worksheet.append(row_values)
        _expand_worksheet_tables_to_data(worksheet)
        workbook.save(output_path)
        return True
    finally:
        workbook.close()


_PNL_CURVE_SHEET = "PnL Curve Features"

_PNL_CURVE_SHEET_COLUMNS: List[str] = [
    "strategy_id",
    "symbol",
    "pnl_curve_fill_concentration",
    "pnl_curve_active_ratio",
    "pnl_curve_time_to_peak",
    "pnl_curve_flatline_pct",
    "pnl_curve_entropy",
    "pnl_curve_shape_class",
    "pnl_curve_points",
    "pnl_curve_raw_json",
]


def _append_pnl_curve_features(
    output_path: Path,
    strategy_id: Any,
    symbol: str,
    features: Any,
) -> None:
    """Append PnL curve features to the dedicated sheet in the workbook.

    Creates the sheet and header row on first use.  Skips duplicate
    strategy_id rows to stay idempotent.
    """
    import openpyxl  # pyright: ignore[reportMissingModuleSource]

    wb = openpyxl.load_workbook(output_path)
    try:
        if _PNL_CURVE_SHEET in wb.sheetnames:
            ws = wb[_PNL_CURVE_SHEET]
        else:
            ws = wb.create_sheet(_PNL_CURVE_SHEET)
            for col_idx, header in enumerate(_PNL_CURVE_SHEET_COLUMNS, start=1):
                ws.cell(row=1, column=col_idx, value=header)

        # Deduplicate by strategy_id
        if strategy_id is not None:
            for existing_row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                if existing_row[0] == strategy_id:
                    logger.info("PnL curve features already exist for strategy_id %s; skipping", strategy_id)
                    return

        feat_dict = features.to_dict()
        row_values = [strategy_id, symbol]
        for col_name in _PNL_CURVE_SHEET_COLUMNS[2:]:
            row_values.append(feat_dict.get(col_name))
        ws.append(row_values)
        _expand_worksheet_tables_to_data(ws)
        wb.save(output_path)
        logger.info("Wrote PnL curve features for %s to '%s' sheet", symbol, _PNL_CURVE_SHEET)
    finally:
        wb.close()




def derive_profit_factor(profits: Sequence[float]) -> Optional[float]:
    """Compute profit factor from realized matched profits."""
    non_zero = [profit for profit in profits if abs(profit) > 0]
    if not non_zero:
        return None

    gross_profits = sum(profit for profit in non_zero if profit > 0)
    gross_losses = abs(sum(profit for profit in non_zero if profit < 0))
    if gross_losses > 0:
        return min(gross_profits / gross_losses, 999.99)
    if gross_profits > 0:
        return 999.99
    return 0.0


def enrich_trade_metrics_from_manual_sections(
    trade_metrics: TradeMetrics,
    matched_profit_events: Sequence[MatchedProfitEvent],
    bot_data: ExtractedBotData,
) -> TradeMetrics:
    """Backfill realized-trade metrics from the manual matched-profit section."""
    if not matched_profit_events:
        return trade_metrics

    profits = [event.profit_usdt for event in matched_profit_events]
    profit_factor = derive_profit_factor(profits)
    gross_pnl = float(sum(profits))
    non_zero_count = len([profit for profit in profits if abs(profit) > 0])

    if trade_metrics.profit_factor in (None, 0.0) and profit_factor is not None:
        trade_metrics.profit_factor = profit_factor
    if trade_metrics.gross_pnl in (None, 0.0):
        trade_metrics.gross_pnl = gross_pnl
    if trade_metrics.win_rate is None and non_zero_count > 0:
        winning = len([profit for profit in profits if profit > 0])
        trade_metrics.win_rate = (winning / non_zero_count) * 100

    if bot_data.total_trades in (None, 0):
        bot_data.total_trades = non_zero_count
    if bot_data.realized_pnl_usdt is None:
        bot_data.realized_pnl_usdt = gross_pnl

    return trade_metrics


def apply_pnl_curve_fallback(
    bot_data: ExtractedBotData,
    trade_metrics: TradeMetrics,
    pnl_curve: Sequence[float],
) -> TradeMetrics:
    """Use the user-provided PnL curve when MTM mark-price data is weak."""
    if len(pnl_curve) < 2:
        return trade_metrics

    curve_mae, curve_mfe = compute_pnl_curve_excursions(list(pnl_curve))
    mtm_unreliable = not trade_metrics.mtm_used_mark_prices

    if (mtm_unreliable or trade_metrics.mae in (None, 0.0)) and curve_mae > 0:
        trade_metrics.mae = curve_mae
    if (mtm_unreliable or trade_metrics.mfe in (None, 0.0)) and curve_mfe > 0:
        trade_metrics.mfe = curve_mfe

    if bot_data.invested_margin_usdt and bot_data.invested_margin_usdt > 0:
        if trade_metrics.mae not in (None, 0.0) and trade_metrics.mae_pct_initial in (None, 0.0):
            trade_metrics.mae_pct_initial = (trade_metrics.mae / bot_data.invested_margin_usdt) * 100
        if trade_metrics.mfe not in (None, 0.0) and trade_metrics.mfe_pct_initial in (None, 0.0):
            trade_metrics.mfe_pct_initial = (trade_metrics.mfe / bot_data.invested_margin_usdt) * 100

    return trade_metrics


def validate_required_fields(bot_data: ExtractedBotData) -> None:
    """Fail fast on missing manual fields the pipeline cannot infer safely."""
    missing: List[str] = []
    if not bot_data.symbol:
        missing.append("symbol")
    if bot_data.strategy_id is None:
        missing.append("strategy_id")
    if bot_data.end_time_utc is None:
        missing.append("end_time_utc")
    if bot_data.start_time_utc is None:
        missing.append("start_time_utc")
    if bot_data.duration_hours is None:
        missing.append("duration_hours")

    if missing:
        raise ValueError(f"Missing required manual fields: {', '.join(missing)}")


def get_required_market_window(bot_data: ExtractedBotData) -> tuple[str, datetime, datetime]:
    """Return the validated market window as non-optional locals."""
    symbol = bot_data.symbol
    start_time_utc = bot_data.start_time_utc
    end_time_utc = bot_data.end_time_utc

    if not symbol or start_time_utc is None or end_time_utc is None:
        raise ValueError("Missing required market window fields after validation.")

    return symbol, start_time_utc, end_time_utc


async def process_manual_text(
    text: str,
    *,
    output_path: Path = Path("data/new_expired_bots.xlsx"),
    trade_csv: Optional[Path] = None,
    txn_csv: Optional[Path] = None,
    results_dir: Optional[Path] = Path("results"),
    verify: bool = True,
    allow_unresolved_coherence: bool = False,
    dry_run: bool = False,
) -> Dict[str, Any]:
    """Parse manual text, fetch market context, and write one workbook row."""
    bot_data, trade_fills, pnl_curve = build_bot_data_from_text(text)
    matched_profit_events = parse_matched_profit_events(text)
    validate_required_fields(bot_data)
    symbol, start_time_utc, end_time_utc = get_required_market_window(bot_data)

    if not dry_run:
        _persist_manual_input_text(text, symbol, bot_data.strategy_id)

    if verify:
        bot_data.display_verification()

    candidate_id: Optional[str] = None
    if bot_data.price_range_low and bot_data.price_range_high and bot_data.grids_count:
        candidate_id = search_candidate_match(
            symbol=symbol,
            grid_lower=bot_data.price_range_low,
            grid_upper=bot_data.price_range_high,
            grids_count=bot_data.grids_count,
            results_dir=results_dir,
        )

    client = BinanceClient()
    try:
        klines = await fetch_historical_klines(client, symbol, start_time_utc)
        session_mark_klines = await fetch_session_mark_klines(
            client,
            symbol,
            start_time_utc,
            end_time_utc,
        )
        indicators = compute_indicators(klines, bot_data)

        funding_fee_events: List[Dict[str, Any]] = []
        if txn_csv and txn_csv.exists():
            tx_df = load_transaction_history(txn_csv)
            funding_fee_events = extract_funding_fee_events_from_transactions(
                tx_df,
                symbol,
                start_time_utc,
                end_time_utc,
            )
            funding_from_transactions = float(
                sum(float(event.get("income", 0.0)) for event in funding_fee_events)
            )
            if funding_from_transactions != 0.0 and bot_data.funding_fee_usdt is None:
                bot_data.funding_fee_usdt = funding_from_transactions

        trade_metrics = TradeMetrics()
        fills_for_coherence: Optional[List[ParsedTradeFill]] = None
        if trade_csv and trade_csv.exists():
            trades_df = load_trade_history(trade_csv)
            filtered_trades = filter_trades_by_symbol_and_time(
                trades_df,
                symbol,
                start_time_utc,
                end_time_utc,
            )
            trade_metrics = compute_trade_metrics(
                filtered_trades,
                mark_klines=session_mark_klines,
                funding_fee_events=funding_fee_events,
                investment_margin_usdt=bot_data.invested_margin_usdt,
            )
        elif trade_fills:
            fills_for_coherence = trade_fills
            fills_df = text_fills_to_dataframe(trade_fills, start_time_utc)
            trade_metrics = compute_trade_metrics(
                fills_df,
                mark_klines=session_mark_klines,
                funding_fee_events=funding_fee_events,
                investment_margin_usdt=bot_data.invested_margin_usdt,
            )

        trade_metrics = enrich_trade_metrics_from_manual_sections(
            trade_metrics,
            matched_profit_events,
            bot_data,
        )
        trade_metrics = apply_pnl_curve_fallback(bot_data, trade_metrics, pnl_curve)

        coherence_warnings = validate_data_coherence(
            bot_data,
            trade_metrics,
            fills_for_coherence,
        )
        repaired_msgs: List[str] = []
        unresolved_msgs: List[str] = []
        if coherence_warnings:
            repaired_msgs, unresolved_msgs = auto_repair_coherence(
                bot_data,
                trade_metrics,
                coherence_warnings,
            )
            for message in repaired_msgs:
                logger.info("COHERENCE REPAIRED: %s", message)
            for message in unresolved_msgs:
                logger.warning("COHERENCE UNRESOLVED: %s", message)

        if unresolved_msgs and not allow_unresolved_coherence:
            raise ValueError(
                "Refusing to write unresolved coherence issues: "
                + "; ".join(unresolved_msgs)
            )

        row = prepare_entry_validation_row(bot_data, indicators, trade_metrics, candidate_id)
        if bot_data.total_trades is not None:
            row["total_trades"] = bot_data.total_trades

        # ── Pattern-profile pre-entry features (F1/F2/F3/F4) ──────────
        # Reuses already-fetched klines. F3 is derived from the first
        # realized funding settlement in (start, start+7h], which is
        # semantically equivalent to the scan-time "expected next-7h
        # carry" since we know what actually happened.
        realized_carry: Optional[float] = None
        try:
            start_ms = int(start_time_utc.timestamp() * 1000)
            horizon_end_ms = int((start_time_utc.timestamp() + 7 * 3600) * 1000)
            carry_records = await client.get_funding_rate(
                symbol=symbol,
                start_time=start_ms,
                end_time=horizon_end_ms,
                limit=10,
            )
            realized_carry = realized_funding_carry_proxy_next_hours(
                carry_records or [],
                entry_time_ms=start_ms,
                horizon_hours=7.0,
            )
        except Exception as exc:
            logger.warning("Realized funding carry fetch failed for %s: %s", symbol, exc)

        try:
            profile_block = compute_profile_feature_block(
                klines_15m=klines.get("15m", pd.DataFrame()),
                klines_1m=klines.get("1m", pd.DataFrame()),
                funding_carry_expected_next_7h=realized_carry,
            )
            for key, value in profile_block.items():
                if value is not None:
                    row[key] = value
        except Exception as exc:
            logger.warning("Profile feature block failed for %s: %s", symbol, exc)

        # ── PnL curve features (stored in separate sheet) ─────────────
        curve_features = extract_pnl_curve_features(pnl_curve) if pnl_curve else None
        if curve_features is not None:
            logger.info(
                "PnL curve features: shape=%s, active=%.1f%%, conc=%.1f%%, peak@%.0f%%",
                curve_features.shape_class,
                curve_features.active_ratio * 100,
                curve_features.fill_concentration * 100,
                curve_features.time_to_peak * 100,
            )

        columns = resolve_workbook_columns(output_path, read_only=dry_run)
        filtered_row = filter_row_to_schema(row, columns)

        if dry_run:
            return filtered_row

        append_row_preserving_schema(output_path, filtered_row, columns)

        # Write PnL curve features to the dedicated sheet
        if curve_features is not None:
            _append_pnl_curve_features(
                output_path,
                strategy_id=row.get("strategy_id"),
                symbol=row.get("symbol", ""),
                features=curve_features,
            )

        return filtered_row
    finally:
        await client.close()


def parse_args() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Extract manual bot text into the current new_expired_bots.xlsx schema."
    )
    parser.add_argument("--text-file", type=Path, help="Path to text file with bot data")
    parser.add_argument(
        "--text-input",
        action="store_true",
        help="Read bot data from stdin (paste, then Ctrl+Z Enter / Ctrl+D)",
    )
    parser.add_argument("--trade-csv", type=Path, help="Optional trade-history CSV")
    parser.add_argument("--txn-csv", type=Path, help="Optional transaction-history CSV")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/new_expired_bots.xlsx"),
        help="Workbook to append without changing its columns",
    )
    parser.add_argument(
        "--results-dir",
        type=Path,
        default=Path("results"),
        help="Directory with deployment_ready CSV files for candidate matching",
    )
    parser.add_argument("--no-verify", action="store_true", help="Skip console verification output")
    parser.add_argument(
        "--allow-unresolved-coherence",
        action="store_true",
        help="Allow writes even when coherence validation cannot fully resolve issues",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Build and validate the row without writing to Excel",
    )
    return parser.parse_args()


def load_text_from_args(args: argparse.Namespace) -> str:
    """Resolve manual text from the selected CLI input."""
    if args.text_file:
        if not args.text_file.exists():
            raise FileNotFoundError(f"Text file not found: {args.text_file}")
        return args.text_file.read_text(encoding="utf-8")

    if args.text_input:
        print("Paste bot data below, then press Ctrl+Z Enter (Windows) or Ctrl+D (Unix):")
        return sys.stdin.read()

    raise ValueError("Provide one of: --text-file or --text-input")


def main() -> None:
    """CLI entry point."""
    args = parse_args()
    try:
        text = load_text_from_args(args)
        row = asyncio.run(
            process_manual_text(
                text,
                output_path=args.output,
                trade_csv=args.trade_csv,
                txn_csv=args.txn_csv,
                results_dir=args.results_dir,
                verify=not args.no_verify,
                allow_unresolved_coherence=args.allow_unresolved_coherence,
                dry_run=args.dry_run,
            )
        )
    except Exception as exc:
        logger.error("%s", exc)
        raise SystemExit(1) from exc

    logger.info(
        "Prepared row for strategy_id=%s symbol=%s output=%s",
        row.get("strategy_id"),
        row.get("symbol"),
        args.output,
    )


if __name__ == "__main__":
    main()
