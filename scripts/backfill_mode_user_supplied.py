"""User-supplied mode backfill (GRIDFIX-001 §1.7 final stage).

The user supplied explicit (strategy_id, mode) assignments for the 30
post-cutoff strategy_ids that lacked TXT-tagged mode after
`backfill_mode_pre_geometric_launch.py` ran. This script consumes the
v2 candidate and emits a v3 candidate with all 30 user-supplied
assignments applied.

Source of truth: USER_SUPPLIED_MODE below — hardcoded in this script
for auditability.

Same zero-data-loss design as the prior two backfills:
  1. Refuse if v3 path exists.
  2. Binary copy v2 -> v3.
  3. Modify ONLY `mode` cells on the target sheet (`General`).
  4. Save v3.
  5. Round-trip verify: every non-mode cell on every sheet must match v2.

Conflict handling: if a user-supplied row already has a different mode
in v2 (which should not happen — v2's missing list was the input to
this script), the script overrides and logs at WARNING.

After the write, the script reports the final mode value_counts and any
strategy_ids still missing.

Run from repo root:

    python scripts/backfill_mode_user_supplied.py
"""
from __future__ import annotations

import logging
import shutil
import sys
import warnings
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

INPUT_PATH = REPO_ROOT / "data" / "new_expired_bots.with_mode_v2.xlsx"
OUTPUT_PATH = REPO_ROOT / "data" / "new_expired_bots.with_mode_v3.xlsx"

MODE_HEADER = "mode"
STRATEGY_ID_HEADER = "strategy_id"

# User-supplied 2026-05-04 (GRIDFIX-001). 30 entries.
USER_SUPPLIED_MODE: Dict[str, str] = {
    "410998243": "arithmetic",
    "411131325": "geometric",
    "411211553": "geometric",
    "411261433": "geometric",
    "411309011": "geometric",
    "410998281": "arithmetic",
    "411144842": "geometric",
    "411212066": "geometric",
    "411263640": "geometric",
    "411333082": "geometric",
    "411039623": "geometric",
    "411145267": "geometric",
    "411221996": "arithmetic",
    "411279994": "geometric",
    "411082794": "geometric",
    "411148910": "geometric",
    "411234224": "geometric",
    "411280132": "arithmetic",
    "411082851": "geometric",
    "411167017": "geometric",
    "411234422": "geometric",
    "411288425": "arithmetic",
    "411110864": "geometric",
    "411187686": "geometric",
    "411256709": "geometric",
    "411192858": "geometric",
    "411125693": "geometric",
    "411187722": "geometric",
    "411256758": "geometric",
    "411198508": "geometric",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger(__name__)


def find_header_column(ws: Worksheet, header_name: str) -> int:
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == header_name:
            return col_idx
    return 0


def apply_user_supplied(
    ws: Worksheet,
    sid_col: int,
    mode_col: int,
) -> tuple[int, int, int, List[str]]:
    """Returns (rows_set, rows_overridden_with_different, rows_already_correct, sids_not_in_sheet)."""
    rows_set = 0
    rows_overridden = 0
    rows_already_correct = 0
    matched_sids: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=False):
        sid_cell = row[sid_col - 1]
        mode_cell = row[mode_col - 1]
        if sid_cell.value is None:
            continue
        sid_str = str(sid_cell.value)
        target_mode = USER_SUPPLIED_MODE.get(sid_str)
        if target_mode is None:
            continue
        matched_sids.add(sid_str)
        current = mode_cell.value
        if current == target_mode:
            rows_already_correct += 1
            continue
        if not isinstance(mode_cell, Cell):
            if isinstance(mode_cell, MergedCell):
                raise RuntimeError(
                    f"refusing to write at row {sid_cell.row}: mode cell is a MergedCell"
                )
            raise RuntimeError(
                f"unexpected cell type at row {sid_cell.row}: {type(mode_cell).__name__}"
            )
        if current is not None:
            rows_overridden += 1
            logger.warning(
                "row %d strategy_id=%s: overriding mode %r -> %r (user-supplied)",
                sid_cell.row, sid_str, current, target_mode,
            )
        mode_cell.value = target_mode
        rows_set += 1

    sids_not_in_sheet = sorted(set(USER_SUPPLIED_MODE.keys()) - matched_sids)
    return rows_set, rows_overridden, rows_already_correct, sids_not_in_sheet


def verify_no_unintended_changes(
    input_path: Path,
    output_path: Path,
    target_sheet_name: str,
    mode_col: int,
) -> List[str]:
    failures: List[str] = []
    wb_in = openpyxl.load_workbook(input_path, read_only=True, data_only=False)
    wb_out = openpyxl.load_workbook(output_path, read_only=True, data_only=False)

    if wb_in.sheetnames != wb_out.sheetnames:
        failures.append(f"sheet list mismatch: input={wb_in.sheetnames} output={wb_out.sheetnames}")
        wb_in.close()
        wb_out.close()
        return failures

    for sheet_name in wb_in.sheetnames:
        ws_i = wb_in[sheet_name]
        ws_o = wb_out[sheet_name]
        max_row_i = ws_i.max_row or 0
        max_row_o = ws_o.max_row or 0
        max_col_i = ws_i.max_column or 0
        max_col_o = ws_o.max_column or 0
        if max_row_i != max_row_o:
            failures.append(
                f"[{sheet_name}] row count differs: input={max_row_i} output={max_row_o}"
            )
        if max_col_i != max_col_o:
            failures.append(
                f"[{sheet_name}] column count differs: input={max_col_i} output={max_col_o}"
            )
        cmp_max_row = min(max_row_i, max_row_o)
        cmp_max_col = min(max_col_i, max_col_o)
        is_target = (sheet_name == target_sheet_name)
        for row_idx in range(1, cmp_max_row + 1):
            for col_idx in range(1, cmp_max_col + 1):
                if is_target and col_idx == mode_col:
                    continue
                v_i = ws_i.cell(row=row_idx, column=col_idx).value
                v_o = ws_o.cell(row=row_idx, column=col_idx).value
                if v_i != v_o:
                    failures.append(
                        f"[{sheet_name}] cell ({row_idx},{col_idx}) differs: "
                        f"input={v_i!r} output={v_o!r}"
                    )
                    if len(failures) >= 50:
                        failures.append("... (truncated at 50 mismatches)")
                        wb_in.close()
                        wb_out.close()
                        return failures

    wb_in.close()
    wb_out.close()
    return failures


def main() -> None:
    if not INPUT_PATH.exists():
        raise SystemExit(
            f"input not found: {INPUT_PATH}\nRun the prior two backfill scripts first."
        )
    if OUTPUT_PATH.exists():
        raise SystemExit(
            f"output already exists: {OUTPUT_PATH}\n"
            "Refusing to overwrite. Inspect or delete the existing v3 candidate, then re-run."
        )

    logger.info("USER_SUPPLIED_MODE entries: %d", len(USER_SUPPLIED_MODE))

    logger.info("--- Binary copy v2 -> v3 ---")
    shutil.copy2(INPUT_PATH, OUTPUT_PATH)
    logger.info("copied %s -> %s (%d bytes)", INPUT_PATH, OUTPUT_PATH, OUTPUT_PATH.stat().st_size)

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        for w in caught:
            logger.warning("openpyxl load warning: %s", w.message)

    target_sheet_name = wb.sheetnames[0]
    ws = wb[target_sheet_name]
    sid_col = find_header_column(ws, STRATEGY_ID_HEADER)
    mode_col = find_header_column(ws, MODE_HEADER)
    if sid_col == 0 or mode_col == 0:
        raise SystemExit(
            f"target sheet '{target_sheet_name}' missing required headers: "
            f"strategy_id={sid_col} mode={mode_col}"
        )
    logger.info(
        "target sheet '%s' columns: strategy_id=%d mode=%d",
        target_sheet_name, sid_col, mode_col,
    )

    logger.info("--- Apply user-supplied mappings ---")
    rows_set, rows_overridden, rows_already_correct, sids_not_in_sheet = apply_user_supplied(
        ws, sid_col, mode_col,
    )
    logger.info(
        "[%s] rule summary: rows_set=%d rows_overridden=%d rows_already_correct=%d "
        "user_sids_not_in_sheet=%d",
        target_sheet_name, rows_set, rows_overridden, rows_already_correct, len(sids_not_in_sheet),
    )
    if sids_not_in_sheet:
        for sid in sids_not_in_sheet:
            logger.warning("user-supplied strategy_id %s NOT FOUND in sheet (no row to write)", sid)

    wb.save(OUTPUT_PATH)
    wb.close()
    logger.info("v3 candidate saved: %s", OUTPUT_PATH)

    logger.info("--- Round-trip verification v2 vs v3 (only mode column may differ) ---")
    failures = verify_no_unintended_changes(INPUT_PATH, OUTPUT_PATH, target_sheet_name, mode_col)
    if failures:
        for msg in failures:
            logger.error(msg)
        raise SystemExit(
            f"VERIFICATION FAILED: {len(failures)} mismatches outside mode column. "
            f"v3 left at {OUTPUT_PATH}; v2 at {INPUT_PATH} untouched."
        )
    logger.info("VERIFICATION PASSED: every non-mode cell matches v2.")

    # Final coverage report
    logger.info("--- Final mode coverage in v3 ---")
    wb_final = openpyxl.load_workbook(OUTPUT_PATH, read_only=True, data_only=False)
    ws_final = wb_final[target_sheet_name]
    sid_col_f = find_header_column(ws_final, STRATEGY_ID_HEADER)
    mode_col_f = find_header_column(ws_final, MODE_HEADER)
    counts: Dict[str, int] = {"arithmetic": 0, "geometric": 0, "NaN/None": 0, "other": 0}
    missing_sids: List[str] = []
    for row in ws_final.iter_rows(min_row=2, values_only=False):
        sid_v = row[sid_col_f - 1].value
        mode_v = row[mode_col_f - 1].value
        if mode_v is None:
            counts["NaN/None"] += 1
            if sid_v is not None:
                missing_sids.append(str(sid_v))
        elif mode_v == "arithmetic":
            counts["arithmetic"] += 1
        elif mode_v == "geometric":
            counts["geometric"] += 1
        else:
            counts["other"] += 1
            logger.warning("unexpected mode value %r at row %d", mode_v, row[0].row)
    wb_final.close()
    logger.info(
        "v3 final mode counts: arithmetic=%d geometric=%d NaN/None=%d other=%d",
        counts["arithmetic"], counts["geometric"], counts["NaN/None"], counts["other"],
    )
    if missing_sids:
        logger.warning("strategy_ids still missing mode in v3: count=%d", len(missing_sids))
        for sid in missing_sids:
            logger.warning("  still missing: strategy_id=%s", sid)
    else:
        logger.info("zero strategy_ids missing mode in v3")

    logger.info(
        "Done. v2 untouched: %s. v3 ready for review: %s.",
        INPUT_PATH, OUTPUT_PATH,
    )


if __name__ == "__main__":
    main()
