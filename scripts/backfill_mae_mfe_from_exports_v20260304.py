"""Backfill MAE/MFE columns in data/new_expired_bots.xlsx from local Binance exports.

Data sources (current architecture):
- Trade exports: data/manual_exports/Trade History*.csv
- Transaction export (funding timestamps): data/manual_exports/Transaction History.csv
- Mark-price path: local price_store parquet, fallback to Binance markPriceKlines API

The script avoids ambiguous attribution by skipping overlapping bot windows
for the same symbol.
"""

from __future__ import annotations

import argparse
import asyncio
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd

from new_bot_data_extractor import (
    compute_trade_metrics,
    extract_funding_fee_events_from_transactions,
)
from neutralgrid.api.binance_client import BinanceClient


@dataclass
class BackfillStats:
    prefilled_rows: int = 0
    prefilled_cells: int = 0
    updated_rows: int = 0
    skipped_rows: int = 0
    skip_reasons: Counter[str] | None = None

    def __post_init__(self) -> None:
        if self.skip_reasons is None:
            self.skip_reasons = Counter()


def _parse_utc(value: Any) -> pd.Timestamp | pd.NaT:
    return pd.to_datetime(value, utc=True, errors="coerce")


def _load_trade_exports(exports_dir: Path) -> pd.DataFrame:
    trade_files = sorted(exports_dir.glob("Trade History*.csv"))
    if not trade_files:
        return pd.DataFrame()
    parts = [pd.read_csv(path) for path in trade_files]
    trades = pd.concat(parts, ignore_index=True)
    time_col = "Time(UTC)" if "Time(UTC)" in trades.columns else "Time"
    if time_col not in trades.columns:
        return pd.DataFrame()
    trades = trades.copy()
    trades["parsed_time"] = pd.to_datetime(trades[time_col], utc=True, errors="coerce")
    return trades


def _load_reconstruction_summary(reconstruction_xlsx: Path) -> pd.DataFrame:
    if not reconstruction_xlsx.exists():
        return pd.DataFrame()
    summary = pd.read_excel(reconstruction_xlsx, sheet_name="Summary")
    if summary.empty:
        return pd.DataFrame()

    mapping = {
        "Strategy ID": "strategy_id",
        "Symbol": "symbol",
        "Status": "status",
        "Start Time": "start_time_utc",
        "End Time": "end_time_utc",
        "Duration (hrs)": "duration_hours",
        "Margin (USDT)": "invested_margin_usdt",
        "Leverage": "leverage",
        "Grids": "grids_count",
        "Range Low": "price_range_low",
        "Range High": "price_range_high",
        "Grid Spacing %": "grid_spacing_pct",
        "Total Profit": "total_profit_usdt",
        "PnL %": "pnl_pct",
        "Trades (replay)": "total_trades",
        "Realized PnL (replay)": "realized_pnl_usdt",
    }
    out = summary.rename(columns=mapping)
    keep_cols = [col for col in mapping.values() if col in out.columns]
    out = out[keep_cols].copy()
    if "strategy_id" in out.columns:
        out["strategy_id"] = pd.to_numeric(out["strategy_id"], errors="coerce").astype("Int64")
    return out


def _merge_missing_from_summary(
    df: pd.DataFrame, summary: pd.DataFrame
) -> tuple[pd.DataFrame, int, int, list[tuple[int, str, Any]]]:
    """Merge missing fields from reconstruction summary.

    Returns
    -------
    tuple of (updated_df, merged_row_count, merged_cell_count, patches)
        ``patches`` is a list of ``(row_idx, col_name, value)`` tuples
        representing every cell that was filled in.
    """
    if summary.empty:
        return df, 0, 0, []

    out = df.copy()
    out["strategy_id"] = pd.to_numeric(out["strategy_id"], errors="coerce").astype("Int64")
    s = summary.copy()
    s["strategy_id"] = pd.to_numeric(s["strategy_id"], errors="coerce").astype("Int64")

    merged_rows: set[int] = set()
    merged_cells = 0
    patches: list[tuple[int, str, Any]] = []

    for idx, row in out.iterrows():
        sid = row.get("strategy_id")
        if pd.isna(sid):
            continue

        matches = s.loc[s["strategy_id"] == sid]
        if matches.empty:
            continue

        symbol = str(row.get("symbol", "") or "")
        if "symbol" in matches.columns and symbol:
            same_symbol = matches.loc[matches["symbol"].astype(str) == symbol]
            if not same_symbol.empty:
                matches = same_symbol

        if len(matches) != 1:
            continue

        src = matches.iloc[0]
        row_updated = False
        for col in src.index:
            if col == "strategy_id" or col not in out.columns:
                continue
            cur = row.get(col)
            new = src.get(col)
            if pd.isna(cur) and pd.notna(new):
                out.at[idx, col] = new
                patches.append((int(idx), col, new))
                row_updated = True
                merged_cells += 1
        if row_updated:
            merged_rows.add(int(idx))

    return out, len(merged_rows), merged_cells, patches


def _overlapping_row_indices(df: pd.DataFrame) -> set[int]:
    overlaps: set[int] = set()
    for _, g in df.groupby("symbol"):
        g = g.sort_values("start_time_utc")
        rows = list(g[["start_time_utc", "end_time_utc"]].itertuples(index=True, name=None))
        for i in range(len(rows)):
            idx_i, s_i, e_i = rows[i]
            if pd.isna(s_i) or pd.isna(e_i):
                continue
            for j in range(i + 1, len(rows)):
                idx_j, s_j, e_j = rows[j]
                if pd.isna(s_j) or pd.isna(e_j):
                    continue
                if s_j > e_i:
                    break
                if s_j <= e_i and e_j >= s_i:
                    overlaps.add(int(idx_i))
                    overlaps.add(int(idx_j))
    return overlaps


def _local_mark_klines(
    price_store_dir: Path,
    symbol: str,
    start: pd.Timestamp,
    end: pd.Timestamp,
) -> list[list[Any]]:
    dates = pd.date_range(start.normalize(), end.normalize(), freq="D", tz="UTC")
    parts: list[pd.DataFrame] = []
    for d in dates:
        path = price_store_dir / symbol / "mark_kline" / "1m" / f"{d.date().isoformat()}.parquet"
        if not path.exists():
            return []
        parts.append(pd.read_parquet(path))
    if not parts:
        return []

    merged = pd.concat(parts, ignore_index=True)
    start_ms = int(start.timestamp() * 1000)
    end_ms = int(end.timestamp() * 1000)
    merged = merged[
        (merged["close_time_ms"] >= start_ms)
        & (merged["close_time_ms"] <= end_ms)
    ]
    if merged.empty:
        return []

    return [
        [
            int(r.open_time_ms),
            str(r.open),
            str(r.high),
            str(r.low),
            str(r.close),
            str(r.volume),
            int(r.close_time_ms),
            "0",
            "0",
            "0",
            "0",
            "0",
        ]
        for r in merged.itertuples(index=False)
    ]


def _filtered_trades_for_window(
    all_trades: pd.DataFrame,
    symbol: str,
    start: pd.Timestamp,
    end: pd.Timestamp,
) -> pd.DataFrame:
    if all_trades.empty:
        return pd.DataFrame()
    symbol_col = "Symbol" if "Symbol" in all_trades.columns else "symbol"
    if symbol_col not in all_trades.columns:
        return pd.DataFrame()
    mask = (
        (all_trades[symbol_col].astype(str) == symbol)
        & (all_trades["parsed_time"] >= start)
        & (all_trades["parsed_time"] <= end)
    )
    return all_trades.loc[mask].copy()


def _apply_patches_to_workbook(
    xlsx_path: Path,
    patches: list[tuple[int, str, Any]],
) -> None:
    """Apply cell-selective patches to the workbook using openpyxl.

    Parameters
    ----------
    xlsx_path : Path
        Path to the Excel workbook.
    patches : list of (row_idx, col_name, value)
        Each ``row_idx`` is the 0-based DataFrame row index.
        ``col_name`` is mapped to a column via the workbook header row.
    """
    import openpyxl

    if not patches:
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Build column-name → 1-based column index map from the header row
    col_map: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            col_map[str(cell.value)] = cell.column

    # Determine the next available column index for any new headers
    next_col = max(col_map.values(), default=0) + 1

    for row_idx, col_name, value in patches:
        col_idx = col_map.get(col_name)
        if col_idx is None:
            # Column not in header yet — add it
            ws.cell(row=1, column=next_col, value=col_name)
            col_map[col_name] = next_col
            col_idx = next_col
            next_col += 1
        # DataFrame row 0 → Excel row 2 (row 1 is the header)
        excel_row = row_idx + 2
        cell_val = value
        # Convert tz-aware datetime to naive for Excel compatibility
        if hasattr(cell_val, "tzinfo") and getattr(cell_val, "tzinfo", None) is not None:
            cell_val = cell_val.replace(tzinfo=None)
        # Convert pandas Timestamp to Python datetime
        if isinstance(cell_val, pd.Timestamp):
            cell_val = cell_val.to_pydatetime().replace(tzinfo=None)
        # Convert bool to int (1/0) for uniform workbook values
        if isinstance(cell_val, bool):
            cell_val = int(cell_val)
        ws.cell(row=excel_row, column=col_idx, value=cell_val)

    wb.save(xlsx_path)
    wb.close()


async def backfill(
    xlsx_path: Path,
    exports_dir: Path,
    price_store_dir: Path,
    tx_filename: str,
    reconstruction_xlsx: Path,
) -> BackfillStats:
    stats = BackfillStats()

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Expired bots file not found: {xlsx_path}")

    trade_df = _load_trade_exports(exports_dir)
    if trade_df.empty:
        raise RuntimeError("No trade export data found (Trade History*.csv)")

    tx_path = exports_dir / tx_filename
    if not tx_path.exists():
        raise FileNotFoundError(f"Transaction export not found: {tx_path}")
    tx_df = pd.read_csv(tx_path)

    df = pd.read_excel(xlsx_path)
    for col in ("mae", "mfe", "mae_pct_initial", "mfe_pct_initial"):
        if col not in df.columns:
            df[col] = pd.NA

    # Collect all cell patches: (row_idx, col_name, value)
    all_patches: list[tuple[int, str, Any]] = []

    summary = _load_reconstruction_summary(reconstruction_xlsx)
    df, prefilled_rows, prefilled_cells, summary_patches = _merge_missing_from_summary(df, summary)
    all_patches.extend(summary_patches)
    stats.prefilled_rows = prefilled_rows
    stats.prefilled_cells = prefilled_cells

    df["start_time_utc"] = pd.to_datetime(df["start_time_utc"], utc=True, errors="coerce")
    df["end_time_utc"] = pd.to_datetime(df["end_time_utc"], utc=True, errors="coerce")

    overlaps = _overlapping_row_indices(df)

    client = BinanceClient()
    try:
        for idx, row in df.iterrows():
            symbol = str(row.get("symbol", ""))
            start = row.get("start_time_utc")
            end = row.get("end_time_utc")

            if int(idx) in overlaps:
                stats.skipped_rows += 1
                stats.skip_reasons["overlap_window"] += 1
                continue
            if pd.isna(start) or pd.isna(end):
                stats.skipped_rows += 1
                stats.skip_reasons["missing_time"] += 1
                continue

            start_ts = _parse_utc(start)
            end_ts = _parse_utc(end)
            if pd.isna(start_ts) or pd.isna(end_ts):
                stats.skipped_rows += 1
                stats.skip_reasons["invalid_time"] += 1
                continue

            row_trades = _filtered_trades_for_window(trade_df, symbol, start_ts, end_ts)
            if row_trades.empty:
                stats.skipped_rows += 1
                stats.skip_reasons["trades_missing"] += 1
                continue

            funding_events = extract_funding_fee_events_from_transactions(
                tx_df=tx_df,
                symbol=symbol,
                start_time=start_ts.to_pydatetime(),
                end_time=end_ts.to_pydatetime(),
            )

            mark_klines = _local_mark_klines(price_store_dir, symbol, start_ts, end_ts)
            if not mark_klines:
                mark_klines = await client.get_mark_price_klines_all(
                    symbol=symbol,
                    interval="1m",
                    start_time=int(start_ts.timestamp() * 1000),
                    end_time=int(end_ts.timestamp() * 1000),
                    include_current=False,
                )
            if not mark_klines:
                stats.skipped_rows += 1
                stats.skip_reasons["mark_missing"] += 1
                continue

            margin_val = pd.to_numeric(pd.Series([row.get("invested_margin_usdt")]), errors="coerce").iloc[0]
            margin = float(margin_val) if pd.notna(margin_val) else None

            metrics = compute_trade_metrics(
                row_trades,
                mark_klines=mark_klines,
                funding_fee_events=funding_events,
                investment_margin_usdt=margin,
            )

            row_i = int(idx)
            all_patches.append((row_i, "mae", metrics.mae))
            all_patches.append((row_i, "mfe", metrics.mfe))
            all_patches.append((row_i, "mae_pct_initial", metrics.mae_pct_initial))
            all_patches.append((row_i, "mfe_pct_initial", metrics.mfe_pct_initial))
            stats.updated_rows += 1
    finally:
        await client.close()

    _apply_patches_to_workbook(xlsx_path, all_patches)
    return stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Backfill MAE/MFE into new_expired_bots.xlsx from exports"
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("data/new_expired_bots.xlsx"),
        help="Path to target XLSX",
    )
    parser.add_argument(
        "--exports-dir",
        type=Path,
        default=Path("data/manual_exports"),
        help="Directory with Trade History*.csv and Transaction History.csv",
    )
    parser.add_argument(
        "--tx-file",
        type=str,
        default="Transaction History.csv",
        help="Transaction history filename under exports-dir",
    )
    parser.add_argument(
        "--price-store",
        type=Path,
        default=Path("data/price_store"),
        help="Local mark kline parquet store root",
    )
    parser.add_argument(
        "--reconstruction-xlsx",
        type=Path,
        default=Path("data/order_book_reconstruction.xlsx"),
        help="Workbook with Summary sheet used to prefill missing bot fields",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    stats = asyncio.run(
        backfill(
            xlsx_path=args.xlsx,
            exports_dir=args.exports_dir,
            price_store_dir=args.price_store,
            tx_filename=args.tx_file,
            reconstruction_xlsx=args.reconstruction_xlsx,
        )
    )
    print("Backfill complete")
    print(f"Prefilled rows: {stats.prefilled_rows}")
    print(f"Prefilled cells: {stats.prefilled_cells}")
    print(f"Updated rows: {stats.updated_rows}")
    print(f"Skipped rows: {stats.skipped_rows}")
    print(f"Skip reasons: {dict(stats.skip_reasons or {})}")


if __name__ == "__main__":
    main()
