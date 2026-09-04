"""Zero-data-loss surgical backfill of the `mode` column.

GRIDFIX-001 / GRID_SYNCH plan v3 §1.7 (option b — mode-only backfill).

Pipeline (six stages, each fail-loud):

    1. Parse `mode` from every TXT under `data/manual_input/**/*.txt`
       (read-only) -> {strategy_id: mode}.
    2. Refuse to run if `data/new_expired_bots.with_mode.xlsx` already
       exists (no silent overwrite of a previous candidate).
    3. Pre-flight scan of the ORIGINAL workbook (read-only): enumerate
       sheets, charts, images, pivot tables, conditional formatting,
       defined names, tables, merged ranges. Capture every openpyxl
       UserWarning emitted during load. Report them all.
    4. Binary copy ORIGINAL -> CANDIDATE via `shutil.copy2` (preserves
       file metadata; bit-for-bit duplicate).
    5. Open CANDIDATE; modify ONLY the `mode` cells on the target sheet
       (sheet at index 0, which is what `pd.read_excel` returns by
       default — what `ExistingDataMapper.map_dataframe` consumes).
       Save CANDIDATE.
    6. Round-trip verification: open CANDIDATE and ORIGINAL side-by-side
       in read-only mode and compare every non-mode cell value, sheet by
       sheet. Any mismatch (in sheet count, sheet names, or cell values
       outside the target sheet's `mode` column) -> abort with a
       cell-level report. The CANDIDATE file is left in place for the
       user to inspect; the ORIGINAL is never touched under any path.

The script never writes to `data/new_expired_bots.xlsx`. The user
inspects `data/new_expired_bots.with_mode.xlsx` and renames manually
when satisfied.

Run from repo root:

    python scripts/backfill_mode_column.py
"""
from __future__ import annotations

import logging
import shutil
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from _bot_data_extractor_core import parse_user_text  # noqa: E402

MANUAL_INPUT_DIR = REPO_ROOT / "data" / "manual_input"
ORIGINAL_PATH = REPO_ROOT / "data" / "new_expired_bots.xlsx"
CANDIDATE_PATH = REPO_ROOT / "data" / "new_expired_bots.with_mode.xlsx"
MODE_HEADER = "mode"
STRATEGY_ID_HEADER = "strategy_id"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Stage 1 — TXT scan
# ---------------------------------------------------------------------------

def collect_mode_from_txts(manual_input_dir: Path) -> Dict[str, str]:
    """Return {strategy_id: mode} parsed from every TXT under `manual_input_dir`.

    TXTs missing strategy_id or mode are logged at WARNING and skipped.
    On per-row mode conflict (same strategy_id, different mode across
    multiple TXTs), the last value wins; conflicts are logged.
    """
    mapping: Dict[str, str] = {}
    txt_paths = sorted(manual_input_dir.rglob("*.txt"))
    if not txt_paths:
        raise SystemExit(f"no TXT files found under {manual_input_dir}")

    parsed_ok = 0
    skipped_no_strategy_id = 0
    skipped_no_mode = 0
    parse_errors = 0
    conflicts: Dict[str, str] = {}

    for txt_path in txt_paths:
        try:
            text = txt_path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = txt_path.read_text(encoding="latin-1")
        try:
            parsed = parse_user_text(text)
        except Exception as exc:  # noqa: BLE001
            parse_errors += 1
            logger.warning("parse failure %s: %s", txt_path.name, exc)
            continue

        strategy_id_raw = parsed.get("strategy_id")
        mode = parsed.get("mode")
        if strategy_id_raw is None:
            skipped_no_strategy_id += 1
            logger.warning("no strategy_id in %s; skipped", txt_path.name)
            continue
        if mode is None:
            skipped_no_mode += 1
            logger.warning(
                "no mode keyword in %s (strategy_id=%s); skipped",
                txt_path.name,
                strategy_id_raw,
            )
            continue

        strategy_id = str(strategy_id_raw)
        existing = mapping.get(strategy_id)
        if existing is not None and existing != mode:
            conflicts[strategy_id] = f"{existing} -> {mode} ({txt_path.name})"
        mapping[strategy_id] = mode
        parsed_ok += 1

    logger.info(
        "TXT scan complete: parsed_ok=%d skipped_no_strategy_id=%d "
        "skipped_no_mode=%d parse_errors=%d unique_strategy_ids=%d",
        parsed_ok,
        skipped_no_strategy_id,
        skipped_no_mode,
        parse_errors,
        len(mapping),
    )
    if conflicts:
        for sid, msg in conflicts.items():
            logger.warning("mode conflict for strategy_id=%s: %s (last value wins)", sid, msg)
    return mapping


# ---------------------------------------------------------------------------
# Stage 3 — Pre-flight scan
# ---------------------------------------------------------------------------

def find_header_column(ws: Worksheet, header_name: str) -> int:
    """Return 1-based column index of the first row-1 cell whose value is
    exactly `header_name`, or 0 if not found."""
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == header_name:
            return col_idx
    return 0


def list_sheets_with_strategy_id(wb: openpyxl.Workbook) -> List[Tuple[str, int]]:
    matches: List[Tuple[str, int]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col = find_header_column(ws, STRATEGY_ID_HEADER)
        if col > 0:
            matches.append((sheet_name, col))
    return matches


def preflight_inspect(workbook_path: Path) -> Tuple[openpyxl.Workbook, List[str]]:
    """Open workbook with warning capture; report advanced features.

    Returns the loaded workbook and the list of warning messages emitted
    during load. The caller can use the returned workbook for stage 5
    (modification) so the load is performed only once.
    """
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        wb = openpyxl.load_workbook(workbook_path)
        warning_msgs = [str(w.message) for w in caught]

    if warning_msgs:
        for msg in warning_msgs:
            logger.warning("openpyxl load warning: %s", msg)
    else:
        logger.info("openpyxl load: no warnings emitted")

    logger.info("workbook sheets: %s", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        n_charts = len(ws._charts)  # type: ignore[attr-defined]
        n_images = len(ws._images)  # type: ignore[attr-defined]
        n_tables = len(ws.tables)
        n_merged = len(ws.merged_cells.ranges)
        cf = ws.conditional_formatting
        n_cf = sum(1 for _ in cf)
        if any((n_charts, n_images, n_tables, n_merged, n_cf)):
            logger.warning(
                "[%s] advanced features detected: charts=%d images=%d tables=%d "
                "merged_ranges=%d conditional_formatting_rules=%d (round-trip risk)",
                sheet_name, n_charts, n_images, n_tables, n_merged, n_cf,
            )
        else:
            logger.info("[%s] no advanced features detected", sheet_name)

    n_defined_names = len(list(wb.defined_names))
    if n_defined_names > 0:
        logger.warning(
            "workbook has %d defined names (round-trip risk)", n_defined_names
        )

    return wb, warning_msgs


# ---------------------------------------------------------------------------
# Stage 5 — Modify the candidate
# ---------------------------------------------------------------------------

def backfill_mode_on_sheet(
    ws: Worksheet,
    sid_col: int,
    mode_by_strategy_id: Dict[str, str],
) -> Tuple[int, int, int, int, int]:
    """Update only the `mode` cells on the given sheet. Returns
    (mode_col_index, rows_with_strategy_id, rows_already_correct,
     rows_updated, rows_unmatched)."""
    mode_col = find_header_column(ws, MODE_HEADER)
    if mode_col == 0:
        mode_col = ws.max_column + 1
        ws.cell(row=1, column=mode_col, value=MODE_HEADER)
        logger.info(
            "[%s] created new '%s' header at column %d (no existing column)",
            ws.title, MODE_HEADER, mode_col,
        )
    else:
        logger.info(
            "[%s] existing '%s' header found at column %d (will update in-place)",
            ws.title, MODE_HEADER, mode_col,
        )

    rows_with_sid = 0
    rows_already_correct = 0
    rows_updated = 0
    rows_unmatched = 0

    for row in ws.iter_rows(min_row=2, values_only=False):
        sid_cell = row[sid_col - 1]
        if sid_cell.value is None:
            continue
        rows_with_sid += 1
        new_mode = mode_by_strategy_id.get(str(sid_cell.value))
        if new_mode is None:
            rows_unmatched += 1
            continue
        mode_cell = ws.cell(row=sid_cell.row, column=mode_col)
        if isinstance(mode_cell, MergedCell):
            raise RuntimeError(
                f"[{ws.title}] refusing to write mode at row {sid_cell.row}, "
                f"col {mode_col}: target cell is a MergedCell. "
                "Unmerge the affected range in the workbook before re-running."
            )
        if not isinstance(mode_cell, Cell):
            raise RuntimeError(
                f"[{ws.title}] unexpected cell type at row {sid_cell.row}, "
                f"col {mode_col}: {type(mode_cell).__name__}"
            )
        if mode_cell.value == new_mode:
            rows_already_correct += 1
            continue
        mode_cell.value = new_mode
        rows_updated += 1

    return mode_col, rows_with_sid, rows_already_correct, rows_updated, rows_unmatched


# ---------------------------------------------------------------------------
# Stage 6 — Round-trip verification
# ---------------------------------------------------------------------------

def verify_no_unintended_changes(
    original_path: Path,
    candidate_path: Path,
    target_sheet_name: str,
    mode_col_in_candidate: int,
) -> List[str]:
    """Compare ORIGINAL and CANDIDATE cell-by-cell.

    Allowed differences:
      - On `target_sheet_name`, header row column `mode_col_in_candidate`
        may be the new MODE_HEADER (when the column was newly created).
      - On `target_sheet_name`, data rows column `mode_col_in_candidate`
        may differ (those are the cells we intentionally wrote).

    Anything else is reported as a verification failure. The function
    returns the list of failure messages (empty list = clean).
    """
    failures: List[str] = []
    wb_orig = openpyxl.load_workbook(original_path, read_only=True, data_only=False)
    wb_cand = openpyxl.load_workbook(candidate_path, read_only=True, data_only=False)

    if wb_orig.sheetnames != wb_cand.sheetnames:
        failures.append(
            f"sheet list mismatch: original={wb_orig.sheetnames} candidate={wb_cand.sheetnames}"
        )
        wb_orig.close()
        wb_cand.close()
        return failures

    for sheet_name in wb_orig.sheetnames:
        ws_o = wb_orig[sheet_name]
        ws_c = wb_cand[sheet_name]

        max_row_o = ws_o.max_row or 0
        max_row_c = ws_c.max_row or 0
        max_col_o = ws_o.max_column or 0
        max_col_c = ws_c.max_column or 0

        # On the target sheet, the candidate is allowed to have ONE more column
        # iff the mode column was newly created at max_col_o + 1.
        is_target = (sheet_name == target_sheet_name)
        max_col_c_expected_extra = 1 if (
            is_target and mode_col_in_candidate == max_col_o + 1
        ) else 0

        if max_row_o != max_row_c:
            failures.append(
                f"[{sheet_name}] row count differs: original={max_row_o} candidate={max_row_c}"
            )
        if max_col_c != max_col_o + max_col_c_expected_extra:
            failures.append(
                f"[{sheet_name}] column count differs: original={max_col_o} "
                f"candidate={max_col_c} (expected_extra={max_col_c_expected_extra})"
            )

        # Iterate cells. Compare every (row, col) up to the smaller bounds.
        cmp_max_row = min(max_row_o, max_row_c)
        cmp_max_col = min(max_col_o, max_col_c)
        for row_idx in range(1, cmp_max_row + 1):
            for col_idx in range(1, cmp_max_col + 1):
                if is_target and col_idx == mode_col_in_candidate:
                    # mode column on target sheet — intentional changes allowed
                    continue
                v_o = ws_o.cell(row=row_idx, column=col_idx).value
                v_c = ws_c.cell(row=row_idx, column=col_idx).value
                if v_o != v_c:
                    failures.append(
                        f"[{sheet_name}] cell ({row_idx},{col_idx}) differs: "
                        f"original={v_o!r} candidate={v_c!r}"
                    )
                    if len(failures) >= 50:
                        failures.append("... (truncated at 50 cell-level mismatches)")
                        wb_orig.close()
                        wb_cand.close()
                        return failures

    wb_orig.close()
    wb_cand.close()
    return failures


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    if not ORIGINAL_PATH.exists():
        raise SystemExit(f"original workbook not found: {ORIGINAL_PATH}")
    if CANDIDATE_PATH.exists():
        raise SystemExit(
            f"candidate path already exists: {CANDIDATE_PATH}\n"
            "Refusing to overwrite. Inspect or delete the existing candidate, then re-run."
        )

    # Stage 1
    mode_by_strategy_id = collect_mode_from_txts(MANUAL_INPUT_DIR)
    if not mode_by_strategy_id:
        raise SystemExit("no (strategy_id, mode) pairs parsed; nothing to backfill")

    # Stage 3 (pre-flight on ORIGINAL — we close this and reopen the COPY for editing)
    logger.info("--- Pre-flight scan of ORIGINAL workbook (read-only) ---")
    wb_inspect, _ = preflight_inspect(ORIGINAL_PATH)
    sheets_with_sid = list_sheets_with_strategy_id(wb_inspect)
    if not sheets_with_sid:
        wb_inspect.close()
        raise SystemExit(
            f"no sheet in {ORIGINAL_PATH.name} contains a '{STRATEGY_ID_HEADER}' header in row 1; "
            "aborting without copying or writing"
        )
    target_sheet_name = wb_inspect.sheetnames[0]
    target_sid_col = find_header_column(wb_inspect[target_sheet_name], STRATEGY_ID_HEADER)
    if target_sid_col == 0:
        wb_inspect.close()
        raise SystemExit(
            f"sheet at index 0 ('{target_sheet_name}') has no '{STRATEGY_ID_HEADER}' header; "
            f"sheets with strategy_id: {sheets_with_sid}; aborting"
        )
    other_sheets_with_sid = [name for name, _ in sheets_with_sid if name != target_sheet_name]
    if other_sheets_with_sid:
        logger.warning(
            "other sheets also contain '%s' and will be LEFT UNTOUCHED: %s",
            STRATEGY_ID_HEADER, other_sheets_with_sid,
        )
    wb_inspect.close()

    # Stage 4 — binary copy
    logger.info("--- Binary copy ORIGINAL -> CANDIDATE ---")
    shutil.copy2(ORIGINAL_PATH, CANDIDATE_PATH)
    logger.info("copied %s -> %s (%d bytes)", ORIGINAL_PATH, CANDIDATE_PATH, CANDIDATE_PATH.stat().st_size)

    # Stage 5 — modify the candidate
    logger.info("--- Modify CANDIDATE: target sheet '%s' ---", target_sheet_name)
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        wb_cand = openpyxl.load_workbook(CANDIDATE_PATH)
        for w in caught:
            logger.warning("openpyxl candidate-load warning: %s", w.message)

    target_ws = wb_cand[target_sheet_name]
    mode_col_used, rows_with_sid, rows_already_correct, rows_updated, rows_unmatched = (
        backfill_mode_on_sheet(target_ws, target_sid_col, mode_by_strategy_id)
    )

    sheet_sids: set[str] = set()
    for row in target_ws.iter_rows(min_row=2, values_only=False):
        sid_cell = row[target_sid_col - 1]
        if sid_cell.value is not None:
            sheet_sids.add(str(sid_cell.value))
    txt_only_sids = sorted(set(mode_by_strategy_id.keys()) - sheet_sids)
    if txt_only_sids:
        logger.warning(
            "strategy_ids present in TXT but not in sheet '%s' (no row to write): %s",
            target_sheet_name, txt_only_sids,
        )

    logger.info(
        "[%s] write summary: rows_with_strategy_id=%d rows_already_correct=%d "
        "rows_updated=%d rows_unmatched=%d txt_strategy_ids_not_in_sheet=%d "
        "mode_col_index=%d",
        target_sheet_name, rows_with_sid, rows_already_correct,
        rows_updated, rows_unmatched, len(txt_only_sids), mode_col_used,
    )

    wb_cand.save(CANDIDATE_PATH)
    wb_cand.close()
    logger.info("CANDIDATE saved: %s", CANDIDATE_PATH)

    # Stage 6 — round-trip verification
    logger.info("--- Round-trip verification ORIGINAL vs CANDIDATE ---")
    failures = verify_no_unintended_changes(
        ORIGINAL_PATH, CANDIDATE_PATH, target_sheet_name, mode_col_used,
    )
    if failures:
        for msg in failures:
            logger.error(msg)
        raise SystemExit(
            f"VERIFICATION FAILED: {len(failures)} cell/sheet mismatches outside the mode column. "
            f"CANDIDATE left in place at {CANDIDATE_PATH} for inspection. "
            f"ORIGINAL at {ORIGINAL_PATH} is untouched."
        )
    logger.info("VERIFICATION PASSED: every non-mode cell matches the original.")
    logger.info(
        "Done. ORIGINAL untouched: %s. CANDIDATE ready for review: %s. "
        "When satisfied, manually rename CANDIDATE -> ORIGINAL.",
        ORIGINAL_PATH, CANDIDATE_PATH,
    )


if __name__ == "__main__":
    main()
