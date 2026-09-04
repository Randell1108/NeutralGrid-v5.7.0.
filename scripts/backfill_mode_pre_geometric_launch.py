"""Pre-geometric-launch arithmetic backfill (GRIDFIX-001 §1.7 follow-up).

Rule (user-supplied 2026-05-04): every grid bot that ENDED before
2026-03-31 12:09:01 UTC predates Binance Futures geometric grid launch
and therefore MUST be arithmetic. Set `mode=arithmetic` on every such
row in the workbook.

This script consumes the v1 candidate produced by
`scripts/backfill_mode_column.py` (which TXT-tags 34 rows) and emits a
v2 candidate that additionally applies the pre-launch arithmetic rule.

Zero-data-loss design (mirrors `backfill_mode_column.py`):
  1. Refuse if output exists.
  2. Binary copy v1 candidate -> v2 candidate.
  3. Modify ONLY `mode` cells on the target sheet (`General`).
  4. Save v2.
  5. Round-trip verify: every non-mode cell on every sheet must match v1.

Conflict handling: if a row already has `mode=geometric` (from TXT) but
ended before the cutoff, the TXT extraction must be wrong (geometric
didn't exist yet). The script overrides geometric -> arithmetic and
logs the conflict at WARNING.

After the write, the script reports the strategy_ids that still have
mode=NaN (rows that ended at-or-after the cutoff and have no
TXT-tagged mode).

Run from repo root:

    python scripts/backfill_mode_pre_geometric_launch.py
"""
from __future__ import annotations

import logging
import shutil
import sys
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Tuple

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

INPUT_PATH = REPO_ROOT / "data" / "new_expired_bots.with_mode.xlsx"
OUTPUT_PATH = REPO_ROOT / "data" / "new_expired_bots.with_mode_v2.xlsx"

CUTOFF_UTC = datetime(2026, 3, 31, 12, 9, 1, tzinfo=timezone.utc)

MODE_HEADER = "mode"
END_TIME_HEADER = "end_time_utc"
STRATEGY_ID_HEADER = "strategy_id"
MODE_ARITHMETIC = "arithmetic"
MODE_GEOMETRIC = "geometric"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger(__name__)


def find_header_column(ws: Worksheet, header_name: str) -> int:
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == header_name:
            return col_idx
    return 0


def coerce_to_utc(value: object) -> datetime | None:
    """Return value as a UTC-aware datetime, or None if value is unparseable."""
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)
    if isinstance(value, str):
        try:
            v = value.replace("Z", "+00:00")
            dt = datetime.fromisoformat(v)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except ValueError:
            return None
    return None


def apply_pre_launch_rule(
    ws: Worksheet,
    end_col: int,
    mode_col: int,
    sid_col: int,
) -> Tuple[int, int, int, int, int, List[str]]:
    """Return (rows_in_scope, rows_already_arithmetic, rows_set_arithmetic,
    rows_overridden_geometric, rows_unparseable_endtime, missing_mode_sids)."""
    rows_in_scope = 0
    rows_already_arithmetic = 0
    rows_set_arithmetic = 0
    rows_overridden_geometric = 0
    rows_unparseable_endtime = 0
    missing_mode_sids: List[str] = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        end_cell = row[end_col - 1]
        mode_cell = row[mode_col - 1]
        sid_cell = row[sid_col - 1]

        end_dt = coerce_to_utc(end_cell.value)
        if end_dt is None:
            if end_cell.value is not None:
                rows_unparseable_endtime += 1
                logger.warning(
                    "row %d: end_time_utc=%r could not be parsed as datetime; mode unchanged",
                    end_cell.row, end_cell.value,
                )
            else:
                if mode_cell.value is None and sid_cell.value is not None:
                    missing_mode_sids.append(str(sid_cell.value))
            continue

        if end_dt < CUTOFF_UTC:
            rows_in_scope += 1
            current = mode_cell.value
            if current == MODE_ARITHMETIC:
                rows_already_arithmetic += 1
                continue
            if not isinstance(mode_cell, Cell):
                if isinstance(mode_cell, MergedCell):
                    raise RuntimeError(
                        f"refusing to write mode at row {end_cell.row}: "
                        f"mode cell is a MergedCell. Unmerge before re-running."
                    )
                raise RuntimeError(
                    f"unexpected cell type at row {end_cell.row}: {type(mode_cell).__name__}"
                )
            if current == MODE_GEOMETRIC:
                rows_overridden_geometric += 1
                logger.warning(
                    "row %d (strategy_id=%s, end_time=%s): mode was geometric but bot "
                    "ended before geometric launch -> overriding to arithmetic",
                    end_cell.row, sid_cell.value, end_dt.isoformat(),
                )
            mode_cell.value = MODE_ARITHMETIC
            rows_set_arithmetic += 1
        else:
            if mode_cell.value is None and sid_cell.value is not None:
                missing_mode_sids.append(str(sid_cell.value))

    return (
        rows_in_scope,
        rows_already_arithmetic,
        rows_set_arithmetic,
        rows_overridden_geometric,
        rows_unparseable_endtime,
        missing_mode_sids,
    )


def verify_no_unintended_changes(
    input_path: Path,
    output_path: Path,
    target_sheet_name: str,
    mode_col: int,
) -> List[str]:
    failures: List[str] = []
    wb_in = openpyxl.load_workbook(input_path, read_only=True, data_only=False)
    wb_out = openpyxl.load_workbook(output_path, read_only=True, data_only=False)

    if wb_in.sheetnames != wb_out.sheetnames:
        failures.append(f"sheet list mismatch: input={wb_in.sheetnames} output={wb_out.sheetnames}")
        wb_in.close()
        wb_out.close()
        return failures

    for sheet_name in wb_in.sheetnames:
        ws_i = wb_in[sheet_name]
        ws_o = wb_out[sheet_name]
        max_row_i = ws_i.max_row or 0
        max_row_o = ws_o.max_row or 0
        max_col_i = ws_i.max_column or 0
        max_col_o = ws_o.max_column or 0
        if max_row_i != max_row_o:
            failures.append(
                f"[{sheet_name}] row count differs: input={max_row_i} output={max_row_o}"
            )
        if max_col_i != max_col_o:
            failures.append(
                f"[{sheet_name}] column count differs: input={max_col_i} output={max_col_o}"
            )
        cmp_max_row = min(max_row_i, max_row_o)
        cmp_max_col = min(max_col_i, max_col_o)
        is_target = (sheet_name == target_sheet_name)
        for row_idx in range(1, cmp_max_row + 1):
            for col_idx in range(1, cmp_max_col + 1):
                if is_target and col_idx == mode_col:
                    continue
                v_i = ws_i.cell(row=row_idx, column=col_idx).value
                v_o = ws_o.cell(row=row_idx, column=col_idx).value
                if v_i != v_o:
                    failures.append(
                        f"[{sheet_name}] cell ({row_idx},{col_idx}) differs: "
                        f"input={v_i!r} output={v_o!r}"
                    )
                    if len(failures) >= 50:
                        failures.append("... (truncated at 50 cell-level mismatches)")
                        wb_in.close()
                        wb_out.close()
                        return failures

    wb_in.close()
    wb_out.close()
    return failures


def main() -> None:
    if not INPUT_PATH.exists():
        raise SystemExit(
            f"input not found: {INPUT_PATH}\n"
            "Run scripts/backfill_mode_column.py first."
        )
    if OUTPUT_PATH.exists():
        raise SystemExit(
            f"output already exists: {OUTPUT_PATH}\n"
            "Refusing to overwrite. Inspect or delete the existing v2 candidate, then re-run."
        )

    logger.info("--- Binary copy v1 -> v2 ---")
    shutil.copy2(INPUT_PATH, OUTPUT_PATH)
    logger.info("copied %s -> %s (%d bytes)", INPUT_PATH, OUTPUT_PATH, OUTPUT_PATH.stat().st_size)

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        for w in caught:
            logger.warning("openpyxl load warning: %s", w.message)

    target_sheet_name = wb.sheetnames[0]
    ws = wb[target_sheet_name]

    sid_col = find_header_column(ws, STRATEGY_ID_HEADER)
    end_col = find_header_column(ws, END_TIME_HEADER)
    mode_col = find_header_column(ws, MODE_HEADER)
    if sid_col == 0 or end_col == 0 or mode_col == 0:
        raise SystemExit(
            f"target sheet '{target_sheet_name}' missing required headers: "
            f"strategy_id={sid_col} end_time_utc={end_col} mode={mode_col}"
        )
    logger.info(
        "target sheet '%s' columns: strategy_id=%d end_time_utc=%d mode=%d",
        target_sheet_name, sid_col, end_col, mode_col,
    )

    logger.info("--- Apply pre-launch rule (cutoff=%s) ---", CUTOFF_UTC.isoformat())
    (
        rows_in_scope,
        rows_already_arithmetic,
        rows_set_arithmetic,
        rows_overridden_geometric,
        rows_unparseable_endtime,
        missing_mode_sids,
    ) = apply_pre_launch_rule(ws, end_col, mode_col, sid_col)

    logger.info(
        "[%s] rule summary: rows_in_scope=%d rows_already_arithmetic=%d "
        "rows_set_arithmetic=%d rows_overridden_geometric=%d "
        "rows_unparseable_endtime=%d",
        target_sheet_name,
        rows_in_scope,
        rows_already_arithmetic,
        rows_set_arithmetic,
        rows_overridden_geometric,
        rows_unparseable_endtime,
    )

    wb.save(OUTPUT_PATH)
    wb.close()
    logger.info("v2 candidate saved: %s", OUTPUT_PATH)

    logger.info("--- Round-trip verification v1 vs v2 (only mode column may differ) ---")
    failures = verify_no_unintended_changes(INPUT_PATH, OUTPUT_PATH, target_sheet_name, mode_col)
    if failures:
        for msg in failures:
            logger.error(msg)
        raise SystemExit(
            f"VERIFICATION FAILED: {len(failures)} mismatches outside the mode column. "
            f"v2 left at {OUTPUT_PATH} for inspection. v1 at {INPUT_PATH} is untouched."
        )
    logger.info("VERIFICATION PASSED: every non-mode cell matches v1.")

    logger.info("--- Strategy IDs still missing mode (post-rule) ---")
    if missing_mode_sids:
        logger.info("count: %d", len(missing_mode_sids))
        for sid in missing_mode_sids:
            logger.info("  missing: strategy_id=%s", sid)
    else:
        logger.info("zero strategy_ids missing mode after rule applied")

    logger.info(
        "Done. v1 untouched: %s. v2 ready for review: %s.",
        INPUT_PATH, OUTPUT_PATH,
    )


if __name__ == "__main__":
    main()
