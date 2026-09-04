"""
Rebuild order book from replay_outputs for bots in new_expired_bots.xlsx.

For each bot, finds the matching replay data by symbol/time range and
reconstructs the order book (buy/sell grid levels) at the initial snapshot.
"""
import csv
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
REPLAY = DATA / "replay_outputs"
INPUT_XLSX = DATA / "new_expired_bots.xlsx"
OUTPUT_XLSX = DATA / "order_book_reconstruction.xlsx"

# ── Style constants ─────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BUY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SELL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
SPREAD_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BOLD = Font(bold=True, size=11)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PRICE_FMT = "#,##0.00000000"
QTY_FMT = "#,##0.00000000"


def to_epoch_ms(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        if val.tzinfo is None:
            val = val.replace(tzinfo=timezone.utc)
        return int(val.timestamp() * 1000)
    if isinstance(val, str):
        for fmt in ["%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"]:
            try:
                dt = datetime.strptime(val, fmt).replace(tzinfo=timezone.utc)
                return int(dt.timestamp() * 1000)
            except ValueError:
                continue
    return None


def ms_to_utc_str(ms):
    return datetime.utcfromtimestamp(ms / 1000).strftime("%Y-%m-%d %H:%M:%S")


def load_csv(path):
    if not path.is_file():
        return []
    with open(path, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def find_best_snapshot_ts(levels, start_ms, end_ms):
    """Return the first in-range snapshot timestamp (initial grid layout)."""
    if not levels:
        return None
    timestamps = sorted(set(int(lv["time_ms"]) for lv in levels))
    if start_ms and end_ms:
        in_range = [ts for ts in timestamps if start_ms <= ts <= end_ms]
        if in_range:
            return in_range[0]
    if start_ms:
        after = [ts for ts in timestamps if ts >= start_ms]
        if after:
            return after[0]
    return timestamps[0]


def styled_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN


def main():
    # ── Load bots ───────────────────────────────────────────────────
    wb_in = openpyxl.load_workbook(str(INPUT_XLSX))
    ws_in = wb_in["Sheet1"]
    hdrs = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    bots = [dict(zip(hdrs, row)) for row in ws_in.iter_rows(min_row=2, max_row=ws_in.max_row, values_only=True)]
    print(f"Loaded {len(bots)} bots from {INPUT_XLSX.name}")

    # ── Pre-load all replay data ────────────────────────────────────
    all_levels, all_snapshots, all_trades = {}, {}, {}
    for sym in os.listdir(REPLAY):
        sym_dir = REPLAY / sym
        if sym_dir.is_dir():
            all_levels[sym] = load_csv(sym_dir / "levels.csv")
            all_snapshots[sym] = load_csv(sym_dir / "snapshots.csv")
            all_trades[sym] = load_csv(sym_dir / "trades.csv")
    print(f"Loaded replay data for {len(all_levels)} symbols")

    # ── Create output workbook ──────────────────────────────────────
    wb = openpyxl.Workbook()

    # ─── Summary sheet ──────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_hdrs = [
        "Strategy ID", "Symbol", "Status", "Start Time", "End Time",
        "Duration (hrs)", "Margin (USDT)", "Leverage", "Grids",
        "Range Low", "Range High", "Grid Spacing %",
        "Total Profit", "PnL %",
        "Buy Levels", "Sell Levels", "Total Open Orders",
        "Best Bid", "Best Ask", "Spread %",
        "Buy Qty", "Sell Qty",
        "Trades (replay)", "Realized PnL (replay)",
        "Replay Data",
    ]
    styled_header_row(ws_sum, 1, sum_hdrs)

    used_sheet_names = set()
    stats = {"total": 0, "matched": 0, "with_levels": 0, "with_trades": 0}

    for i, bot in enumerate(bots):
        stats["total"] += 1
        symbol = bot["symbol"]
        sid = bot["strategy_id"]
        start_ms = to_epoch_ms(bot.get("start_time_utc"))
        end_ms = to_epoch_ms(bot.get("end_time_utc"))

        levels = all_levels.get(symbol, [])
        snapshots = all_snapshots.get(symbol, [])
        trades = all_trades.get(symbol, [])
        has_data = bool(levels)
        if has_data:
            stats["matched"] += 1

        best_ts = find_best_snapshot_ts(levels, start_ms, end_ms)
        ts_levels = [lv for lv in levels if best_ts and int(lv["time_ms"]) == best_ts]
        ts_snap = [s for s in snapshots if best_ts and int(s["time_ms"]) == best_ts]

        buy_lvls = sorted(
            [lv for lv in ts_levels if lv["side"] == "BUY" and float(lv["price"]) > 0],
            key=lambda x: -float(x["price"]),
        )
        sell_lvls = sorted(
            [lv for lv in ts_levels if lv["side"] == "SELL" and float(lv["price"]) > 0],
            key=lambda x: float(x["price"]),
        )
        if buy_lvls or sell_lvls:
            stats["with_levels"] += 1

        snap = ts_snap[0] if ts_snap else None
        best_bid = float(snap["best_bid"]) if snap and snap.get("best_bid") else None
        best_ask = float(snap["best_ask"]) if snap and snap.get("best_ask") else None
        spread_pct = None
        if best_bid and best_ask and best_bid > 0:
            spread_pct = (best_ask - best_bid) / best_bid

        bot_trades = []
        if trades and start_ms and end_ms:
            bot_trades = [t for t in trades if start_ms <= int(t["t_trade_ms"]) <= end_ms]
        if bot_trades:
            stats["with_trades"] += 1

        total_realized = sum(float(t.get("realized_profit", 0) or 0) for t in bot_trades)
        total_buy_qty = sum(float(lv["total_orig_qty"]) for lv in buy_lvls)
        total_sell_qty = sum(float(lv["total_orig_qty"]) for lv in sell_lvls)

        # Summary row
        r = i + 2
        vals = [
            sid, symbol, bot.get("status"),
            str(bot.get("start_time_utc", "")), str(bot.get("end_time_utc", "")),
            bot.get("duration_hours"), bot.get("invested_margin_usdt"),
            bot.get("leverage"), bot.get("grids_count"),
            bot.get("price_range_low"), bot.get("price_range_high"),
            bot.get("grid_spacing_pct"),
            bot.get("total_profit_usdt"), bot.get("pnl_pct"),
            len(buy_lvls), len(sell_lvls), len(buy_lvls) + len(sell_lvls),
            best_bid, best_ask,
            round(spread_pct * 100, 4) if spread_pct else None,
            total_buy_qty, total_sell_qty,
            len(bot_trades),
            round(total_realized, 4) if bot_trades else None,
            "Yes" if has_data else "No",
        ]
        for col, v in enumerate(vals, 1):
            c = ws_sum.cell(row=r, column=col, value=v)
            c.border = THIN

        # ─── Per-bot sheet ──────────────────────────────────────────
        sname = f"{symbol}_{sid}"[:31]
        if sname in used_sheet_names:
            sname = f"{symbol}_{str(sid)[-6:]}"[:31]
        used_sheet_names.add(sname)
        ws = wb.create_sheet(title=sname)

        # Header info
        info = [
            ("Symbol", symbol, "Strategy ID", sid, "Status", bot.get("status")),
            ("Start", str(bot.get("start_time_utc", "")), "End", str(bot.get("end_time_utc", "")),
             "Duration (hrs)", bot.get("duration_hours")),
            ("Range Low", bot.get("price_range_low"), "Range High", bot.get("price_range_high"),
             "Grids", bot.get("grids_count")),
            ("Margin", bot.get("invested_margin_usdt"), "Leverage", bot.get("leverage"),
             "Snapshot", ms_to_utc_str(best_ts) if best_ts else "No data"),
        ]
        for row_idx, row_data in enumerate(info, 1):
            for col_idx in range(0, len(row_data), 2):
                ws.cell(row=row_idx, column=col_idx + 1, value=row_data[col_idx]).font = BOLD
                ws.cell(row=row_idx, column=col_idx + 2, value=row_data[col_idx + 1])

        # Order book table
        ob_row = 6
        ob_hdrs = ["Level #", "Side", "Price", "Qty", "Orders", "Notional (USDT)"]
        styled_header_row(ws, ob_row, ob_hdrs)

        cr = ob_row + 1

        # Sells: highest at top -> lowest near spread
        for j, lv in enumerate(reversed(sell_lvls), 1):
            p = float(lv["price"])
            q = float(lv["total_orig_qty"])
            oc = int(lv["order_count"])
            row_vals = [j, "SELL", p, q, oc, p * q]
            for col, v in enumerate(row_vals, 1):
                c = ws.cell(row=cr, column=col, value=v)
                c.fill = SELL_FILL
                c.border = THIN
                if col == 3:
                    c.number_format = PRICE_FMT
                elif col in (4, 6):
                    c.number_format = QTY_FMT
            cr += 1

        # Spread row
        if best_bid and best_ask:
            for col in range(1, 7):
                ws.cell(row=cr, column=col).fill = SPREAD_FILL
                ws.cell(row=cr, column=col).border = THIN
            ws.cell(row=cr, column=2, value="-- SPREAD --").font = Font(bold=True, italic=True)
            c = ws.cell(row=cr, column=3, value=best_ask - best_bid)
            c.number_format = PRICE_FMT
            c.font = Font(bold=True)
            if spread_pct:
                ws.cell(row=cr, column=4, value=f"{spread_pct*100:.4f}%")
            cr += 1

        # Buys: highest near spread -> lowest at bottom
        for j, lv in enumerate(buy_lvls, 1):
            p = float(lv["price"])
            q = float(lv["total_orig_qty"])
            oc = int(lv["order_count"])
            row_vals = [j, "BUY", p, q, oc, p * q]
            for col, v in enumerate(row_vals, 1):
                c = ws.cell(row=cr, column=col, value=v)
                c.fill = BUY_FILL
                c.border = THIN
                if col == 3:
                    c.number_format = PRICE_FMT
                elif col in (4, 6):
                    c.number_format = QTY_FMT
            cr += 1

        # Stats
        cr += 1
        ws.cell(row=cr, column=1, value="Book Stats").font = Font(bold=True, size=12)
        cr += 1
        book_stats = [
            ("Buy Levels", len(buy_lvls)),
            ("Sell Levels", len(sell_lvls)),
            ("Total Orders", len(buy_lvls) + len(sell_lvls)),
            ("Total Buy Qty", round(total_buy_qty, 8)),
            ("Total Sell Qty", round(total_sell_qty, 8)),
            ("Best Bid", best_bid),
            ("Best Ask", best_ask),
            ("Spread %", f"{spread_pct*100:.4f}%" if spread_pct else "N/A"),
            ("Replay Trades", len(bot_trades)),
            ("Realized PnL", round(total_realized, 4) if bot_trades else "N/A"),
        ]
        for label, val in book_stats:
            ws.cell(row=cr, column=1, value=label).font = BOLD
            ws.cell(row=cr, column=2, value=val)
            cr += 1

        # Trade history
        if bot_trades:
            cr += 1
            ws.cell(row=cr, column=1, value="Trade History").font = Font(bold=True, size=12)
            cr += 1
            t_hdrs = ["Time (UTC)", "Order ID", "Side", "Price", "Qty", "Fee", "Realized Profit"]
            styled_header_row(ws, cr, t_hdrs)
            cr += 1
            for t in bot_trades:
                t_vals = [
                    t.get("time_utc", ""),
                    int(t["order_id"]),
                    t["side"],
                    float(t["price"]),
                    float(t["qty"]),
                    float(t.get("fee", 0) or 0),
                    float(t.get("realized_profit", 0) or 0),
                ]
                fill = BUY_FILL if t["side"] == "BUY" else SELL_FILL
                for col, v in enumerate(t_vals, 1):
                    c = ws.cell(row=cr, column=col, value=v)
                    c.fill = fill
                    c.border = THIN
                cr += 1

        # Column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20

    # Finalize summary
    ws_sum.freeze_panes = "A2"
    for col in range(1, len(sum_hdrs) + 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = 17

    wb.save(str(OUTPUT_XLSX))
    print(f"\nResults:")
    print(f"  Total bots:          {stats['total']}")
    print(f"  With replay data:    {stats['matched']}")
    print(f"  With order levels:   {stats['with_levels']}")
    print(f"  With trade history:  {stats['with_trades']}")
    print(f"\nSaved to: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
