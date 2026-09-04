"""Tests for the focused manual-input extractor."""

from __future__ import annotations

import os
import shutil
import sys
from pathlib import Path
from datetime import datetime, timezone
from types import SimpleNamespace
from uuid import uuid4

import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

import new_bot_data_extractor as nbde
from new_bot_data_extractor import TechnicalIndicators


SAMPLE_TEXT = """STEEMUSDT neutral grid bot data:
-    Est. Liq. Price (Long): 0.044707
-    Est. Liq. Price (Short): 0.086145

Canceled
Isolated
Neutral 10x
Time Ended
2026-03-08 14:30:57
,
Duration
17h 21m

PNL
Total Profit (USDT)
4.99 (1.86%)
Invested Margin (USDT)
269.00
Matched Profit (USDT)
6.26 (2.32%)
Unmatched PNL (USDT)
0.00 (0.00%)
Funding Fee (USDT)
-1.27 (-0.47%)
Annualized Yield
1,175.18%

PnL Curve:
18.    0.00000
19.    0.00000
20.    0.00000
21.    0.00000
22.    0.00000
23.    1.869701
24.    4.990441
25.    4.990441
26.    4.990441
27.    4.990441
28.    4.990441
29.    4.990441
30.    4.990441
31.    4.990441
32.    4.990441
33.    4.990441
34.    4.990441

Grid Details
Realized Profit
+4.99 USDT
Mode
Arithmetic
Price Range
0.05506 - 0.06236 USDT
Number of Grids
6
Profit Per Grid
1.94% - 2.16%
Invested Margin
269.00 USDT
Qty Per Order
5236 STEEM
Initial Leverage
10x
Grid Start Price
0.05932 USDT
Initial Buy Qty
--
Position Margin
--
Margin used by open orders
--
Total Current Margin
--
Strategy Number
410514600
Time Created
2026-03-07 21:07:49
Advanced (Optional)
Trailing Up
Disabled
Trailing Down
Disabled
Trigger Price
0.059450 USDT
Mark
Stop Loss
-32.28 USDT ≈ -12% (Last)
Take Profit
+40.35 USDT ≈ +15% (Last)
Close all positions on stop
Enabled
Close all positions on TP/SL stop
Enabled

History
Total Matched Profit
6.26 USDT
24H Matched Trades
0
Total Matched Trades
1

Trade list
Time    Matched Profits
2026-03-08 03:22:28    6.26114598 USDT
Time    Side    Order Type    Avg. Price    Executed    Total    Fee
2026-03-08 03:22:28    Buy    Limit    0.059920 USDT    5236 STEEM    313.74112000 USDT    0.06274822 USDT
2026-03-08 02:42:56    Sell    Limit    0.061140 USDT    5236 STEEM    320.12904000 USDT    0.0640258 USDT
"""


class DummyClient:
    async def close(self) -> None:
        return None


@pytest.fixture
def workspace_tmp_dir() -> Path:
    root = Path("tests") / ".tmp_new_bot_data_extractor"
    root.mkdir(parents=True, exist_ok=True)
    case_dir = root / uuid4().hex
    case_dir.mkdir()
    try:
        yield case_dir
    finally:
        shutil.rmtree(case_dir, ignore_errors=True)


def test_parse_matched_profit_events_extracts_manual_history():
    events = nbde.parse_matched_profit_events(SAMPLE_TEXT)
    assert len(events) == 1
    assert events[0].profit_usdt == pytest.approx(6.26114598)
    assert events[0].timestamp == datetime(2026, 3, 8, 8, 22, 28, tzinfo=timezone.utc)


def test_manual_ui_bot_timestamps_are_converted_from_local_time_to_utc():
    parsed = nbde.parse_user_text(
        """DOGEUSDT
Ended on 2026-05-14 10:00:00
Time Created
2026-05-14 08:00:00
"""
    )

    assert parsed["start_time_utc"] == datetime(2026, 5, 14, 13, 0, 0, tzinfo=timezone.utc)
    assert parsed["end_time_utc"] == datetime(2026, 5, 14, 15, 0, 0, tzinfo=timezone.utc)


def test_manual_ocr_end_timestamp_is_converted_from_local_time_to_utc():
    parsed = nbde.parse_screenshot_text(
        """DOGEUSDT
Ended on: 2026-05-14 10:00:00
"""
    )

    assert parsed["end_time_utc"] == datetime(2026, 5, 14, 15, 0, 0, tzinfo=timezone.utc)


def test_manual_trade_fill_timestamps_are_converted_from_local_time_to_utc():
    fills = nbde.parse_trade_fills_from_text(
        """Time    Side    Order Type    Avg. Price    Executed    Total    Fee
2026-05-14 08:00:00    Buy    Limit    0.170000 USDT    100 DOGE    17.00000000 USDT    0.00340000 USDT
"""
    )

    assert len(fills) == 1
    assert fills[0].timestamp == datetime(2026, 5, 14, 13, 0, 0, tzinfo=timezone.utc)


def test_parse_matched_profit_events_skips_unmatched_blocks_and_keeps_later_matches():
    text = """Trade list
Time    Matched Profits
2026-03-08 23:44:37    Unmatched Sell
Time    Side    Order Type    Avg. Price    Executed    Total    Fee
2026-03-08 23:44:37    Sell    Limit    0.371100 USDT    1405.0 KAITO    521.39550000 USDT    0.10427909 USDT

2026-03-08 22:39:14    5.69362201 USDT
Time    Side    Order Type    Avg. Price    Executed    Total    Fee
2026-03-08 22:39:14    Buy    Limit    0.366900 USDT    1405.0 KAITO    515.49450000 USDT    0.10309889 USDT
2026-03-08 22:12:28    Sell    Limit    0.371100 USDT    1405.0 KAITO    521.39550000 USDT    0.1042791 USDT

2026-03-08 20:15:46    5.69834280 USDT
"""
    events = nbde.parse_matched_profit_events(text)
    assert [event.profit_usdt for event in events] == pytest.approx([5.69362201, 5.6983428])


def test_filter_row_to_schema_drops_unapproved_keys():
    row = {column: idx for idx, column in enumerate(nbde.DEFAULT_WORKBOOK_COLUMNS)}
    row["coherence_ok"] = 0
    filtered = nbde.filter_row_to_schema(row, nbde.DEFAULT_WORKBOOK_COLUMNS)
    assert list(filtered.keys()) == nbde.DEFAULT_WORKBOOK_COLUMNS
    assert "coherence_ok" not in filtered


def test_pnl_curve_sheet_columns_defined():
    assert nbde._PNL_CURVE_SHEET == "PnL Curve Features"
    assert nbde._PNL_CURVE_SHEET_COLUMNS[0] == "strategy_id"
    assert nbde._PNL_CURVE_SHEET_COLUMNS[1] == "symbol"
    assert len(nbde._PNL_CURVE_SHEET_COLUMNS) == 10  # 2 keys + 8 features


def test_append_row_expands_canonical_table_reference(workspace_tmp_dir):
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.table import Table

    output_path = workspace_tmp_dir / "new_expired_bots.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "General"
    columns = ["strategy_id", "symbol"]
    worksheet.append(columns)
    worksheet.append([1, "BTCUSDT"])
    worksheet.add_table(Table(displayName="GeneralCanonicalTable", ref="A1:B2"))
    workbook.save(output_path)
    workbook.close()

    appended = nbde.append_row_preserving_schema(
        output_path,
        {"strategy_id": 2, "symbol": "ETHUSDT"},
        columns,
    )

    assert appended is True
    verified = load_workbook(output_path)
    try:
        table = verified["General"].tables["GeneralCanonicalTable"]
        assert table.ref == "A1:B3"
    finally:
        verified.close()


@pytest.mark.asyncio
async def test_process_manual_text_writes_current_schema(workspace_tmp_dir, monkeypatch):
    output_path = workspace_tmp_dir / "new_expired_bots.xlsx"

    async def fake_fetch_historical_klines(*args, **kwargs):
        empty = pd.DataFrame()
        return {"1h": empty, "15m": empty, "5m": empty, "1m": empty}

    async def fake_fetch_session_mark_klines(*args, **kwargs):
        return []

    monkeypatch.setattr(nbde, "BinanceClient", DummyClient)
    monkeypatch.setattr(nbde, "fetch_historical_klines", fake_fetch_historical_klines)
    monkeypatch.setattr(nbde, "fetch_session_mark_klines", fake_fetch_session_mark_klines)
    monkeypatch.setattr(nbde, "compute_indicators", lambda *args, **kwargs: TechnicalIndicators())
    monkeypatch.setattr(nbde, "search_candidate_match", lambda **kwargs: None)

    row = await nbde.process_manual_text(
        SAMPLE_TEXT,
        output_path=output_path,
        verify=False,
    )

    workbook_df = pd.read_excel(output_path)
    assert list(workbook_df.columns) == nbde.DEFAULT_WORKBOOK_COLUMNS
    assert "coherence_ok" not in workbook_df.columns
    assert row["strategy_id"] == 410514600
    assert workbook_df.loc[0, "strategy_id"] == 410514600
    assert workbook_df.loc[0, "profit_factor"] == pytest.approx(999.99)

    curve_df = pd.read_excel(output_path, sheet_name=nbde._PNL_CURVE_SHEET)
    assert list(curve_df.columns) == nbde._PNL_CURVE_SHEET_COLUMNS
    assert curve_df.loc[0, "strategy_id"] == 410514600
    assert curve_df.loc[0, "symbol"] == "STEEMUSDT"
    assert curve_df.loc[0, "pnl_curve_points"] == 17


@pytest.mark.asyncio
async def test_process_manual_text_preserves_explicit_zero_total_trades(workspace_tmp_dir, monkeypatch):
    output_path = workspace_tmp_dir / "new_expired_bots.xlsx"
    text = SAMPLE_TEXT.replace("Total Matched Trades\n1", "Total Matched Trades\n0")
    text = text.replace("2026-03-08 03:22:28    6.26114598 USDT", "2026-03-08 03:22:28    Unmatched Sell")

    async def fake_fetch_historical_klines(*args, **kwargs):
        empty = pd.DataFrame()
        return {"1h": empty, "15m": empty, "5m": empty, "1m": empty}

    async def fake_fetch_session_mark_klines(*args, **kwargs):
        return []

    monkeypatch.setattr(nbde, "BinanceClient", DummyClient)
    monkeypatch.setattr(nbde, "fetch_historical_klines", fake_fetch_historical_klines)
    monkeypatch.setattr(nbde, "fetch_session_mark_klines", fake_fetch_session_mark_klines)
    monkeypatch.setattr(nbde, "compute_indicators", lambda *args, **kwargs: TechnicalIndicators())
    monkeypatch.setattr(nbde, "search_candidate_match", lambda **kwargs: None)

    row = await nbde.process_manual_text(
        text,
        output_path=output_path,
        verify=False,
    )

    assert row["total_trades"] == 0


@pytest.mark.asyncio
async def test_process_manual_text_rejects_unresolved_coherence(workspace_tmp_dir, monkeypatch):
    output_path = workspace_tmp_dir / "new_expired_bots.xlsx"

    async def fake_fetch_historical_klines(*args, **kwargs):
        empty = pd.DataFrame()
        return {"1h": empty, "15m": empty, "5m": empty, "1m": empty}

    async def fake_fetch_session_mark_klines(*args, **kwargs):
        return []

    monkeypatch.setattr(nbde, "BinanceClient", DummyClient)
    monkeypatch.setattr(nbde, "fetch_historical_klines", fake_fetch_historical_klines)
    monkeypatch.setattr(nbde, "fetch_session_mark_klines", fake_fetch_session_mark_klines)
    monkeypatch.setattr(nbde, "compute_indicators", lambda *args, **kwargs: TechnicalIndicators())
    monkeypatch.setattr(nbde, "search_candidate_match", lambda **kwargs: None)
    monkeypatch.setattr(nbde, "validate_data_coherence", lambda *args, **kwargs: ["Grid spacing non-positive: -1.0%"])
    monkeypatch.setattr(
        nbde,
        "auto_repair_coherence",
        lambda *args, **kwargs: ([], ["Grid spacing non-positive: -1.0%"]),
    )

    with pytest.raises(ValueError, match="Refusing to write unresolved coherence issues"):
        await nbde.process_manual_text(
            SAMPLE_TEXT,
            output_path=output_path,
            verify=False,
        )

    assert not output_path.exists()
