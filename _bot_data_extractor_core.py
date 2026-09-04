"""Private compatibility core for new_bot_data_extractor.py.

This module preserves the legacy extractor helper surface without requiring the
public runtime to import the deprecated wrapper module.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import math
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import numpy as np
import pandas as pd

# Project imports
from neutralgrid.core.config import get_config
from neutralgrid.api.binance_client import BinanceClient
from neutralgrid.indicators.technical import (
    calc_adx,
    calc_atr,
    calc_bollinger_bands,
    calc_ema,
    calc_ema_slope,
    calc_rsi,
    calc_vwap,
    count_ema_crosses,
    count_vwap_crosses,
    detect_trend_structure,
)
from neutralgrid.metrics.excursion_mtm_v20260304 import calculate_mtm_excursions

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

_MANUAL_UI_TIMEZONE = "America/Lima"


def parse_manual_ui_datetime_to_utc(value: str) -> datetime:
    """Parse a Binance UI/manual timestamp entered in local time and return UTC."""

    local_dt = datetime.strptime(value.strip(), "%Y-%m-%d %H:%M:%S")
    try:
        local_tz = ZoneInfo(_MANUAL_UI_TIMEZONE)
    except ZoneInfoNotFoundError:
        local_tz = timezone(timedelta(hours=-5))
    return local_dt.replace(tzinfo=local_tz).astimezone(timezone.utc)


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class ExtractedBotData:
    """Data extracted from bot screenshots/text/JSON."""
    strategy_id: Optional[int] = None
    symbol: Optional[str] = None
    strategy_type: str = "grid"
    status: Optional[str] = None
    end_time_utc: Optional[datetime] = None
    duration_hours: Optional[float] = None
    invested_margin_usdt: Optional[float] = None
    leverage: Optional[int] = None
    grids_count: Optional[int] = None
    price_range_low: Optional[float] = None
    price_range_high: Optional[float] = None
    total_profit_usdt: Optional[float] = None
    pnl_pct: Optional[float] = None
    realized_pnl_usdt: Optional[float] = None
    unrealized_pnl_usdt: Optional[float] = None
    funding_fee_usdt: Optional[float] = None
    total_trades: Optional[int] = None
    apy_pct: Optional[float] = None
    grid_start_price: Optional[float] = None
    qty_per_order: Optional[float] = None
    mode: Optional[str] = None  # "arithmetic" or "geometric"
    profit_per_grid_pct_low: Optional[float] = None
    profit_per_grid_pct_high: Optional[float] = None
    liq_price_long: Optional[float] = None
    liq_price_short: Optional[float] = None
    trigger_price: Optional[float] = None

    # Computed fields
    start_time_utc: Optional[datetime] = None
    grid_spacing_pct: Optional[float] = None

    def compute_derived_fields(self):
        """Compute derived fields from extracted data."""
        # Calculate start_time (only if not explicitly provided)
        if not self.start_time_utc and self.end_time_utc and self.duration_hours:
            self.start_time_utc = self.end_time_utc - timedelta(hours=self.duration_hours)

        # Calculate grid spacing — MODE-AWARE
        if self.price_range_high and self.price_range_low and self.grids_count and self.grids_count > 1:
            low = self.price_range_low
            high = self.price_range_high
            n = self.grids_count - 1
            if self.mode == "geometric" and low > 0:
                self.grid_spacing_pct = (math.pow(high / low, 1.0 / n) - 1.0) * 100
            else:
                # Arithmetic (default)
                range_size = high - low
                self.grid_spacing_pct = (range_size / low * 100) / n

    def display_verification(self):
        """Display extracted data for user verification."""
        print("\n" + "=" * 60)
        print("EXTRACTED DATA - PLEASE VERIFY")
        print("=" * 60)
        print(f"  Strategy ID:       {self.strategy_id}")
        print(f"  Symbol:            {self.symbol}")
        print(f"  Status:            {self.status}")
        print(f"  Mode:              {self.mode}")
        print(f"  Invested Margin:   {self.invested_margin_usdt} USDT")
        print(f"  Leverage:          {self.leverage}x")
        print(f"  Number of Grids:   {self.grids_count}")
        print(f"  Price Range:       {self.price_range_low} - {self.price_range_high}")
        print(f"  Grid Spacing:      {self.grid_spacing_pct:.4f}%" if self.grid_spacing_pct else "  Grid Spacing:      N/A")
        print(f"  Grid Start Price:  {self.grid_start_price}")
        print(f"  Qty per Order:     {self.qty_per_order}")
        print(f"  Duration:          {self.duration_hours:.2f} hours" if self.duration_hours else "  Duration:          N/A")
        print(f"  End Time:          {self.end_time_utc}")
        print(f"  Start Time:        {self.start_time_utc}")
        print(f"  Total Profit:      {self.total_profit_usdt} USDT ({self.pnl_pct}%)")
        print(f"  Matched Profit:    {self.realized_pnl_usdt} USDT")
        print(f"  Unmatched PnL:     {self.unrealized_pnl_usdt} USDT")
        print(f"  Funding Fee:       {self.funding_fee_usdt} USDT")
        print(f"  Total Trades:      {self.total_trades}")
        print(f"  APY:               {self.apy_pct}%")
        print(f"  Liq Long:          {self.liq_price_long}")
        print(f"  Liq Short:         {self.liq_price_short}")
        print(f"  Trigger Price:     {self.trigger_price}")
        print("=" * 60)


@dataclass
class TechnicalIndicators:
    """Technical indicators computed at bot start time."""
    adx_1h: Optional[float] = None
    adx_15m: Optional[float] = None
    adx_5m: Optional[float] = None
    rsi_15m: Optional[float] = None
    ema_slope_1h: Optional[float] = None
    ema_crosses_5m: Optional[int] = None
    vwap_crosses_5m: Optional[int] = None
    range_size_pct: Optional[float] = None
    bb_width: Optional[float] = None
    trend_structure: Optional[str] = None
    atr_15m: Optional[float] = None
    atr_pct_15m: Optional[float] = None
    last_price: Optional[float] = None


@dataclass
class TradeMetrics:
    """Metrics computed from trade history."""
    executed_qty: float = 0.0
    commission_usdt: float = 0.0
    maker_count: int = 0
    taker_count: int = 0
    notional_completed: float = 0.0
    profit_factor: Optional[float] = None
    win_rate: Optional[float] = None
    mae: Optional[float] = None
    mfe: Optional[float] = None
    mae_pct_initial: Optional[float] = None
    mfe_pct_initial: Optional[float] = None
    gross_pnl: Optional[float] = None
    mtm_used_mark_prices: bool = False


@dataclass
class ParsedTradeFill:
    """A single parsed trade fill from user text."""
    side: str  # "BUY" or "SELL"
    price: float
    quantity: float
    fee: float = 0.0
    realized_profit: float = 0.0
    maker: bool = True
    timestamp: Optional[datetime] = None


# =============================================================================
# TEXT INPUT PARSING (PRIMARY INPUT MODE)
# =============================================================================

def parse_user_text(text: str) -> Dict[str, Any]:
    """Parse user-pasted text containing bot data.

    Handles standard Binance Futures grid bot copy-paste format.
    Returns a dict suitable for constructing ExtractedBotData.
    """
    data: Dict[str, Any] = {}

    # --- Symbol (first 15 lines) ---
    head = "\n".join(text.splitlines()[:15])
    sym_m = re.search(r'\b(\S{1,15}USDT)\b', head)
    if sym_m:
        data["symbol"] = sym_m.group(1)

    # --- Strategy ID / Strategy Number ---
    sid_m = re.search(r'Strategy\s*(?:ID|Number)[:\s]*(\d{9,12})', text, re.IGNORECASE)
    if sid_m:
        data["strategy_id"] = int(sid_m.group(1))

    # --- Status ---
    if re.search(r'\bExpired\b', text, re.IGNORECASE):
        data["status"] = "expired"
    elif re.search(r'\bCancel(?:l)?ed\b', text, re.IGNORECASE):
        data["status"] = "cancelled"
    elif re.search(r'\bEnded\b', text, re.IGNORECASE):
        data["status"] = "ended"

    # --- Mode ---
    if re.search(r'\bArithmetic\b', text, re.IGNORECASE):
        data["mode"] = "arithmetic"
    elif re.search(r'\bGeometric\b', text, re.IGNORECASE):
        data["mode"] = "geometric"

    # --- End time ---
    end_m = re.search(
        r'(?:Ended?\s*(?:on|Time)?)[:\s]*(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2})',
        text, re.IGNORECASE,
    )
    if end_m:
        data["end_time_utc"] = parse_manual_ui_datetime_to_utc(
            f"{end_m.group(1)} {end_m.group(2)}"
        )

    # --- Start time / Time Created ---
    start_m = re.search(
        r'(?:Time\s*Created|Created\s*(?:on|at)?)[:\s]*(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2})',
        text, re.IGNORECASE,
    )
    if start_m:
        data["start_time_utc"] = parse_manual_ui_datetime_to_utc(
            f"{start_m.group(1)} {start_m.group(2)}"
        )

    # --- Duration ---
    dur_m = re.search(r'(\d+)\s*[dD][\s:]*(\d+)\s*[hH][\s:]*(\d+)\s*[mM]', text)
    if dur_m:
        days, hours, mins = int(dur_m.group(1)), int(dur_m.group(2)), int(dur_m.group(3))
        data["duration_hours"] = days * 24 + hours + mins / 60
    else:
        dur_alt = re.search(r'(\d+)\s*[dD]\s+(\d{1,2}):(\d{2}):(\d{2})', text)
        if dur_alt:
            d, h, m, s = (int(dur_alt.group(i)) for i in range(1, 5))
            data["duration_hours"] = d * 24 + h + m / 60 + s / 3600
        else:
            # Hours + minutes only (no days), e.g. "17h 21m". The Binance
            # copy-paste UI also omits the trailing ``m`` in its two-line
            # Duration block (for example, ``Duration\\n6h 23``).
            dur_hm = re.search(
                r'Duration\s*[:\n]\s*(\d+)\s*[hH][\s:]*(\d+)\b',
                text,
                re.IGNORECASE,
            )
            if not dur_hm:
                dur_hm = re.search(r'(\d+)\s*[hH][\s:]*(\d+)\s*[mM]', text)
            if dur_hm:
                data["duration_hours"] = int(dur_hm.group(1)) + int(dur_hm.group(2)) / 60
            else:
                dur_mins = re.search(r'Duration[:\s]*(\d+)\s*[mM]\b', text, re.IGNORECASE)
                if dur_mins:
                    data["duration_hours"] = int(dur_mins.group(1)) / 60
                else:
                    dur_hrs = re.search(r'Duration[:\s]*([\d.]+)\s*(?:hours?|hrs?)', text, re.IGNORECASE)
                    if dur_hrs:
                        data["duration_hours"] = float(dur_hrs.group(1))

    # --- Invested Margin ---
    margin_m = re.search(
        r'\bInvested\s*Margin\b(?:\s*\(USDT\))?[:\s]*([\d,.]+)\s*(?:USDT)?',
        text,
        re.IGNORECASE,
    )
    if not margin_m:
        margin_m = re.search(
            r'^\s*Margin\s*[:\s]+([\d,.]+)\s*(?:USDT)?',
            text,
            re.IGNORECASE | re.MULTILINE,
        )
    if margin_m:
        data["invested_margin_usdt"] = float(margin_m.group(1).replace(",", ""))

    # --- Leverage ---
    lev_m = re.search(r'(?:Initial\s*)?Leverage[:\s]*(\d+)\s*[xX]', text, re.IGNORECASE)
    if lev_m:
        data["leverage"] = int(lev_m.group(1))

    # --- Number of Grids ---
    grid_m = re.search(r'(?:Number\s*of\s*)?Grids[:\s]*(\d+)', text, re.IGNORECASE)
    if grid_m:
        data["grids_count"] = int(grid_m.group(1))

    # --- Price Range ---
    range_m = re.search(
        r'Price\s*Range[:\s]*([\d.]+)\s*[-\u2013\u2014]\s*([\d.]+)', text, re.IGNORECASE
    )
    if range_m:
        data["price_range_low"] = float(range_m.group(1))
        data["price_range_high"] = float(range_m.group(2))

    # --- Grid Start Price ---
    start_p = re.search(r'Grid\s*Start\s*Price[:\s]*([\d.]+)', text, re.IGNORECASE)
    if start_p:
        data["grid_start_price"] = float(start_p.group(1))

    # --- Qty/Order or Qty Per Order ---
    qty_m = re.search(r'Qty\s*(?:/|Per)\s*Order[:\s]*([\d,.]+)', text, re.IGNORECASE)
    if qty_m:
        data["qty_per_order"] = float(qty_m.group(1).replace(",", ""))

    # --- Profit/Grid or Profit Per Grid ---
    pg_m = re.search(
        r'\bProfit\s*(?:/|Per)\s*Grid\b[:\s]*([\d.]+)\s*%(?:\s*[-\u2013\u2014]\s*([\d.]+)\s*%)?',
        text, re.IGNORECASE,
    )
    if pg_m:
        data["profit_per_grid_pct_low"] = float(pg_m.group(1))
        if pg_m.group(2):
            data["profit_per_grid_pct_high"] = float(pg_m.group(2))

    # --- Total Profit (strict anchor avoiding "Total Matched") ---
    # Handles both "Total Profit: 4.99 USDT" and "Total Profit (USDT)\n4.99 (1.86%)"
    tp_m = re.search(
        r'\bTotal\s*Profit\b(?!\s*Matched)[^\d+-]*([-+]?\d[\d,.]*)\s*(?:USDT)?',
        text, re.IGNORECASE,
    )
    if tp_m:
        data["total_profit_usdt"] = float(tp_m.group(1).replace(",", ""))

    # --- PnL % ---
    pnl_m = re.search(
        r'Total\s*Profit.*?\(\s*([-+]?\d[\d,.]*)\s*%\s*\)', text, re.IGNORECASE | re.DOTALL
    )
    if not pnl_m:
        pnl_m = re.search(r'\b(?:PnL|ROI)\b[:\s]*([-+]?\d[\d,.]*)\s*%', text, re.IGNORECASE)
    if pnl_m:
        data["pnl_pct"] = float(pnl_m.group(1).replace(",", ""))

    # --- Matched Profit ---
    # Handles "Matched Profit: 6.26 USDT" and "Matched Profit (USDT)\n6.26 (2.32%)"
    mp_m = re.search(
        r'\bMatched\s*Profit\b[^\d+-]*([-+]?\d[\d,.]*)(?:\s*USDT)?', text, re.IGNORECASE
    )
    if mp_m:
        data["realized_pnl_usdt"] = float(mp_m.group(1).replace(",", ""))

    # --- Unmatched PNL ---
    # Handles "Unmatched PNL: 0.00" and "Unmatched PNL (USDT)\n0.00 (0.00%)"
    up_m = re.search(r'Unmatched\s*PNL?[^\d+-]*([-+]?\d[\d,.]*)', text, re.IGNORECASE)
    if up_m:
        data["unrealized_pnl_usdt"] = float(up_m.group(1).replace(",", ""))

    # --- Funding Fee ---
    # Handles "Funding Fee: -1.27" and "Funding Fee (USDT)\n-1.27 (-0.47%)"
    ff_m = re.search(r'Funding\s*Fee[^\d+-]*([-+]?\d[\d,.]*)', text, re.IGNORECASE)
    if ff_m:
        data["funding_fee_usdt"] = float(ff_m.group(1).replace(",", ""))

    # --- Total Trades ---
    tt_m = re.search(r'Total\s*(?:Matched\s*)?Trades[:\s]*(\d+)', text, re.IGNORECASE)
    if tt_m:
        data["total_trades"] = int(tt_m.group(1))

    # --- APY ---
    apy_m = re.search(
        r'(?:Annualized\s*)?(?:Yield|APY)[:\s]*([-\d,.]+)\s*%?', text, re.IGNORECASE
    )
    if apy_m:
        data["apy_pct"] = float(apy_m.group(1).replace(",", ""))

    # --- Liquidation Prices ---
    # Handles: "Liq Long", "Liq. Price (Long)", "Est. Liq. Price (Long)", "Liquidation Price (Long)"
    liq_l = re.search(
        r'Liq(?:uidation)?\.?\s*(?:Price\s*)?\(?Long\)?[:\s]*([\d.]+)', text, re.IGNORECASE
    )
    if liq_l:
        data["liq_price_long"] = float(liq_l.group(1))

    liq_s = re.search(
        r'Liq(?:uidation)?\.?\s*(?:Price\s*)?\(?Short\)?[:\s]*([\d.]+)', text, re.IGNORECASE
    )
    if liq_s:
        data["liq_price_short"] = float(liq_s.group(1))

    # --- Trigger Price ---
    trig_m = re.search(r'Trigger\s*Price[:\s]*([\d.]+)', text, re.IGNORECASE)
    if trig_m:
        data["trigger_price"] = float(trig_m.group(1))

    return data


def parse_trade_fills_from_text(text: str) -> List[ParsedTradeFill]:
    """Extract individual trade fills from user-pasted text.

    Handles formats:
      Buy  180.50 USDT  0.10  Fee: 0.01  Realized: 1.80
      BUY 180.50 qty=0.10 fee=0.01 pnl=0.00
      1. Buy 180.50 0.058 0.01808 0.00
    """
    fills: List[ParsedTradeFill] = []

    # Flexible pattern: [optional_num] Side Price [USDT] Qty [Fee] [Realized]
    pattern = re.compile(
        r'(?:^|\n)\s*(?:\d+[.)]\s*)?'
        r'(Buy|Sell|BUY|SELL)\s+'
        r'([\d.]+)\s*(?:USDT)?\s+'
        r'([\d.]+)'
        r'(?:\s+(?:Fee[:\s=]*)?([\d.]+)(?:\s*USDT)?)?'
        r'(?:\s+(?:(?:Realized|PnL|Profit)[:\s=]*)?([-\d.]+)(?:\s*USDT)?)?',
        re.IGNORECASE | re.MULTILINE,
    )

    for m in pattern.finditer(text):
        side = m.group(1).upper()
        price = float(m.group(2))
        qty = float(m.group(3))
        fee = float(m.group(4)) if m.group(4) else 0.0
        realized = float(m.group(5)) if m.group(5) else 0.0
        if price > 0 and qty > 0:
            fills.append(ParsedTradeFill(
                side=side,
                price=price,
                quantity=qty,
                fee=fee,
                realized_profit=realized,
            ))

    if not fills:
        # Binance tab-separated format:
        # 2026-03-08 03:22:28    Buy    Limit    0.059920 USDT    5236 STEEM    313.74 USDT    0.0627 USDT
        tab_pattern = re.compile(
            r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+'
            r'(Buy|Sell)\s+'
            r'(?:Limit|Market)\s+'
            r'([\d.]+)\s*USDT\s+'
            r'([\d,.]+)\s*\w+\s+'        # qty + asset
            r'([\d,.]+)\s*USDT\s+'        # total notional
            r'([\d,.]+)\s*USDT',          # fee
            re.IGNORECASE,
        )
        for m in tab_pattern.finditer(text):
            ts_str = m.group(1).strip()
            side = m.group(2).upper()
            price = float(m.group(3))
            qty = float(m.group(4).replace(",", ""))
            fee = float(m.group(6).replace(",", ""))
            try:
                ts = parse_manual_ui_datetime_to_utc(ts_str)
            except ValueError:
                ts = None
            if price > 0 and qty > 0:
                fills.append(ParsedTradeFill(
                    side=side,
                    price=price,
                    quantity=qty,
                    fee=fee,
                    realized_profit=0.0,
                    timestamp=ts,
                ))

    return fills


def parse_pnl_curve_from_text(text: str) -> List[float]:
    """Extract cumulative PnL curve values from a Binance manual dump.

    Numbered entries are accepted anywhere in a contiguous curve block.  Bare
    numeric values are accepted only after an explicit ``PnL Curve`` heading,
    which is how the current Binance copy-paste UI renders the series.
    """
    if not text:
        return []

    lines = text.splitlines()
    value_pattern = re.compile(
        r'^\s*(\d+)[.):]?\s+([-+]?[\d,.]+)\s*(?:USDT|usdt)?\s*$'
    )
    bare_value_pattern = re.compile(r'^\s*([-+]?[\d,.]+)\s*(?:USDT|usdt)?\s*$')

    start_index = 0
    for i, line in enumerate(lines):
        if "pnl curve" in line.lower():
            start_index = i + 1
            break

    values: List[float] = []
    started = False
    for line in lines[start_index:]:
        stripped = line.strip()
        if not stripped:
            if started:
                continue
            continue

        match = value_pattern.match(stripped)
        if match:
            try:
                values.append(float(match.group(2).replace(",", "")))
                started = True
            except ValueError:
                continue
            continue

        bare_match = bare_value_pattern.match(stripped)
        if start_index > 0 and bare_match:
            try:
                values.append(float(bare_match.group(1).replace(",", "")))
                started = True
            except ValueError:
                continue
            continue

        if started:
            break

    if values or start_index > 0:
        return values

    for line in lines:
        stripped = line.strip()
        if not stripped:
            if started:
                break
            continue

        match = value_pattern.match(stripped)
        if match:
            try:
                values.append(float(match.group(2).replace(",", "")))
                started = True
            except ValueError:
                continue
            continue

        if started:
            break

    return values


def compute_pnl_curve_excursions(pnl_values: List[float]) -> Tuple[float, float]:
    """Compute MAE and MFE from a cumulative PnL curve.

    MAE = max peak-to-trough drawdown (positive value).
    MFE = max trough-to-peak runup (positive value).
    """
    if len(pnl_values) < 2:
        return 0.0, 0.0

    peak = pnl_values[0]
    trough = pnl_values[0]
    max_drawdown = 0.0
    max_runup = 0.0

    for val in pnl_values[1:]:
        if val > peak:
            peak = val
        drawdown = peak - val
        if drawdown > max_drawdown:
            max_drawdown = drawdown

        if val < trough:
            trough = val
        runup = val - trough
        if runup > max_runup:
            max_runup = runup

    return max_drawdown, max_runup


def text_fills_to_dataframe(
    fills: List[ParsedTradeFill],
    start_time: Optional[datetime] = None,
) -> pd.DataFrame:
    """Convert parsed trade fills to a DataFrame compatible with compute_trade_metrics."""
    if not fills:
        return pd.DataFrame()

    base_ts = start_time or datetime.now(timezone.utc)
    rows = []
    for i, f in enumerate(fills):
        ts = f.timestamp or (base_ts + timedelta(minutes=i))
        rows.append({
            "Side": f.side.capitalize(),
            "Price": f"{f.price} USDT",
            "Quantity": str(f.quantity),
            "Fee": f"{f.fee} USDT",
            "Realized Profit": f"{f.realized_profit} USDT",
            "Maker": "TRUE" if f.maker else "FALSE",
            "parsed_time": ts,
        })

    return pd.DataFrame(rows)


def build_bot_data_from_text(
    text: str,
) -> Tuple[ExtractedBotData, List[ParsedTradeFill], List[float]]:
    """Parse user text into (ExtractedBotData, trade_fills, pnl_curve)."""
    parsed = parse_user_text(text)
    fills = parse_trade_fills_from_text(text)
    pnl_curve = parse_pnl_curve_from_text(text)

    valid_fields = {
        "strategy_id", "symbol", "strategy_type", "status",
        "start_time_utc", "end_time_utc",
        "duration_hours", "invested_margin_usdt", "leverage", "grids_count",
        "price_range_low", "price_range_high", "total_profit_usdt", "pnl_pct",
        "realized_pnl_usdt", "unrealized_pnl_usdt", "funding_fee_usdt",
        "total_trades", "apy_pct", "grid_start_price", "qty_per_order", "mode",
        "profit_per_grid_pct_low", "profit_per_grid_pct_high",
        "liq_price_long", "liq_price_short", "trigger_price",
    }
    filtered = {k: v for k, v in parsed.items() if k in valid_fields}
    bot_data = ExtractedBotData(**filtered)
    bot_data.compute_derived_fields()

    logger.info(
        "Parsed text: symbol=%s, fills=%d, pnl_points=%d",
        bot_data.symbol, len(fills), len(pnl_curve),
    )
    return bot_data, fills, pnl_curve


# =============================================================================
# OCR EXTRACTION (LEGACY)
# =============================================================================

def try_ocr_extraction(image_path: Path) -> Optional[str]:
    """Attempt OCR extraction using available libraries."""
    try:
        import pytesseract  # type: ignore[reportMissingImports]
        from PIL import Image

        img = Image.open(image_path)
        text = pytesseract.image_to_string(img)
        logger.info(f"OCR extracted {len(text)} characters from {image_path.name}")
        return text
    except ImportError:
        logger.warning("pytesseract not available, OCR disabled")
        return None
    except Exception as e:
        logger.error(f"OCR failed for {image_path}: {e}")
        return None


def parse_screenshot_text(text: str) -> Dict[str, Any]:
    """Parse OCR text from a bot screenshot to extract key fields."""
    data = {}

    strategy_match = re.search(r'Strategy\s*ID[:\s]*(\d{9,12})', text, re.IGNORECASE)
    if strategy_match:
        data['strategy_id'] = int(strategy_match.group(1))

    head = "\n".join(text.splitlines()[:10])
    symbol_match = re.search(r'\b(\S{2,15}USDT)\b', head)
    if symbol_match:
        data['symbol'] = symbol_match.group(1)

    if re.search(r'\bExpired\b', text, re.IGNORECASE):
        data['status'] = 'expired'
    elif re.search(r'\bCancell?ed\b', text, re.IGNORECASE):
        data['status'] = 'cancelled'
    elif re.search(r'\bEnded\b', text, re.IGNORECASE):
        data['status'] = 'ended'

    if re.search(r'\bArithmetic\b', text, re.IGNORECASE):
        data['mode'] = 'arithmetic'
    elif re.search(r'\bGeometric\b', text, re.IGNORECASE):
        data['mode'] = 'geometric'

    end_match = re.search(r'Ended\s*on[:\s]*(\d{4}-\d{2}-\d{2}\s*\d{2}:\d{2}:\d{2})', text, re.IGNORECASE)
    if end_match:
        data['end_time_utc'] = parse_manual_ui_datetime_to_utc(end_match.group(1).strip())

    duration_match = re.search(r'(\d+)\s*[dD][\s:]*(\d+)\s*[hH][\s:]*(\d+)\s*[mM]', text)
    if duration_match:
        days = int(duration_match.group(1))
        hours = int(duration_match.group(2))
        mins = int(duration_match.group(3))
        data['duration_hours'] = days * 24 + hours + mins / 60
    else:
        duration_alt = re.search(r'(\d+)\s*[dD]\s+(\d{1,2}):(\d{2}):(\d{2})', text)
        if duration_alt:
            days = int(duration_alt.group(1))
            hh = int(duration_alt.group(2))
            mm = int(duration_alt.group(3))
            ss = int(duration_alt.group(4))
            data['duration_hours'] = days * 24 + hh + mm / 60 + ss / 3600
        else:
            duration_hm = re.search(r'(\d+)\s*[hH][\s:]*(\d+)\s*[mM]', text)
            if duration_hm:
                data['duration_hours'] = int(duration_hm.group(1)) + int(duration_hm.group(2)) / 60

    margin_match = re.search(r'Invested\s*Margin[:\s]*([\d,.]+)\s*(?:USDT)?', text, re.IGNORECASE)
    if margin_match:
        data['invested_margin_usdt'] = float(margin_match.group(1).replace(',', ''))

    leverage_match = re.search(r'(?:Initial\s*)?Leverage[:\s]*(\d+)\s*[xX]', text, re.IGNORECASE)
    if leverage_match:
        data['leverage'] = int(leverage_match.group(1))

    grids_match = re.search(r'(?:Number\s*of\s*)?Grids[:\s]*(\d+)', text, re.IGNORECASE)
    if grids_match:
        data['grids_count'] = int(grids_match.group(1))

    range_match = re.search(r'Price\s*Range[:\s]*([\d.]+)\s*[-\u2013]\s*([\d.]+)', text, re.IGNORECASE)
    if range_match:
        data['price_range_low'] = float(range_match.group(1))
        data['price_range_high'] = float(range_match.group(2))

    start_price_match = re.search(r'Grid\s*Start\s*Price[:\s]*([\d.]+)', text, re.IGNORECASE)
    if start_price_match:
        data['grid_start_price'] = float(start_price_match.group(1))

    qty_match = re.search(r'Qty/Order[:\s]*([\d,.]+)', text, re.IGNORECASE)
    if qty_match:
        data['qty_per_order'] = float(qty_match.group(1).replace(',', ''))

    profit_grid_match = re.search(
        r'\bProfit/Grid\b[:\s]*([\d.]+)\s*%(?:\s*[-\u2013\u2014]\s*([\d.]+)\s*%)?',
        text, re.IGNORECASE
    )
    if profit_grid_match:
        data['profit_per_grid_pct_low'] = float(profit_grid_match.group(1))
        if profit_grid_match.group(2):
            data['profit_per_grid_pct_high'] = float(profit_grid_match.group(2))

    profit_match = re.search(r'\bTotal\s*Profit\b(?!\s*Matched)[^\d-]*([-\d,.]+)\s*USDT', text, re.IGNORECASE)
    if profit_match:
        data['total_profit_usdt'] = float(profit_match.group(1).replace(',', ''))

    pnl_match = re.search(r'Total\s*Profit.*?\(\s*([-\d,.]+)\s*%\s*\)', text, re.IGNORECASE | re.DOTALL)
    if not pnl_match:
        pnl_match = re.search(r'\b(?:PnL|ROI)\b[:\s]*([-\d,.]+)\s*%', text, re.IGNORECASE)
    if pnl_match:
        data['pnl_pct'] = float(pnl_match.group(1).replace(',', ''))

    matched_match = re.search(r'\bMatched\s*Profit\b[:\s]*([-\d,.]+)(?:\s*USDT)?', text, re.IGNORECASE)
    if matched_match:
        data['realized_pnl_usdt'] = float(matched_match.group(1).replace(',', ''))

    unmatched_match = re.search(r'Unmatched\s*PNL?[:\s]*([-\d,.]+)', text, re.IGNORECASE)
    if unmatched_match:
        data['unrealized_pnl_usdt'] = float(unmatched_match.group(1).replace(',', ''))

    funding_match = re.search(r'Funding\s*Fee[:\s]*([-\d,.]+)', text, re.IGNORECASE)
    if funding_match:
        data['funding_fee_usdt'] = float(funding_match.group(1).replace(',', ''))

    trades_match = re.search(r'Total\s*(?:Matched\s*)?Trades[:\s]*(\d+)', text, re.IGNORECASE)
    if trades_match:
        data['total_trades'] = int(trades_match.group(1))

    apy_match = re.search(r'(?:Annualized\s*)?(?:Yield|APY)[:\s]*([-\d,.]+)\s*%?', text, re.IGNORECASE)
    if apy_match:
        data['apy_pct'] = float(apy_match.group(1).replace(',', ''))

    # Trigger price
    trigger_match = re.search(r'Trigger\s*Price[:\s]*([\d,.]+)', text, re.IGNORECASE)
    if trigger_match:
        data['trigger_price'] = float(trigger_match.group(1).replace(',', ''))

    # Liquidation prices (long / short)
    liq_long_match = re.search(
        r'Liq(?:uidation)?\s*(?:Price\s*)?\(?Long\)?[:\s]*([\d,.]+)', text, re.IGNORECASE
    )
    if liq_long_match:
        data['liq_price_long'] = float(liq_long_match.group(1).replace(',', ''))

    liq_short_match = re.search(
        r'Liq(?:uidation)?\s*(?:Price\s*)?\(?Short\)?[:\s]*([\d,.]+)', text, re.IGNORECASE
    )
    if liq_short_match:
        data['liq_price_short'] = float(liq_short_match.group(1).replace(',', ''))

    return data


def extract_from_images(image_paths: List[Path]) -> ExtractedBotData:
    """Extract bot data from multiple screenshot images."""
    combined_data = {}

    for img_path in image_paths:
        if not img_path.exists():
            logger.warning(f"Image not found: {img_path}")
            continue

        text = try_ocr_extraction(img_path)
        if text:
            parsed = parse_screenshot_text(text)
            existing_sid = combined_data.get('strategy_id')
            new_sid = parsed.get('strategy_id')
            if existing_sid and new_sid and existing_sid != new_sid:
                logger.warning(f"Skipping {img_path.name}: strategy_id mismatch {new_sid} != {existing_sid}")
                continue
            for key, value in parsed.items():
                if value is None:
                    continue
                if combined_data.get(key) is None:
                    combined_data[key] = value

    bot_data = ExtractedBotData(
        strategy_id=combined_data.get('strategy_id'),
        symbol=combined_data.get('symbol'),
        status=combined_data.get('status'),
        end_time_utc=combined_data.get('end_time_utc'),
        duration_hours=combined_data.get('duration_hours'),
        invested_margin_usdt=combined_data.get('invested_margin_usdt'),
        leverage=combined_data.get('leverage'),
        grids_count=combined_data.get('grids_count'),
        price_range_low=combined_data.get('price_range_low'),
        price_range_high=combined_data.get('price_range_high'),
        total_profit_usdt=combined_data.get('total_profit_usdt'),
        pnl_pct=combined_data.get('pnl_pct'),
        realized_pnl_usdt=combined_data.get('realized_pnl_usdt'),
        unrealized_pnl_usdt=combined_data.get('unrealized_pnl_usdt'),
        funding_fee_usdt=combined_data.get('funding_fee_usdt'),
        total_trades=combined_data.get('total_trades'),
        apy_pct=combined_data.get('apy_pct'),
        grid_start_price=combined_data.get('grid_start_price'),
        qty_per_order=combined_data.get('qty_per_order'),
        mode=combined_data.get('mode'),
        profit_per_grid_pct_low=combined_data.get('profit_per_grid_pct_low'),
        profit_per_grid_pct_high=combined_data.get('profit_per_grid_pct_high'),
        liq_price_long=combined_data.get('liq_price_long'),
        liq_price_short=combined_data.get('liq_price_short'),
        trigger_price=combined_data.get('trigger_price'),
    )

    bot_data.compute_derived_fields()
    return bot_data


def load_manual_input(json_path: Path) -> ExtractedBotData:
    """Load manually extracted data from a JSON file."""
    with open(json_path, 'r') as f:
        data = json.load(f)

    if isinstance(data.get('end_time_utc'), str):
        data['end_time_utc'] = datetime.strptime(data['end_time_utc'], '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)

    valid_fields = {
        'strategy_id', 'symbol', 'strategy_type', 'status', 'end_time_utc',
        'duration_hours', 'invested_margin_usdt', 'leverage', 'grids_count',
        'price_range_low', 'price_range_high', 'total_profit_usdt', 'pnl_pct',
        'realized_pnl_usdt', 'unrealized_pnl_usdt', 'funding_fee_usdt',
        'total_trades', 'apy_pct', 'grid_start_price', 'qty_per_order', 'mode',
        'profit_per_grid_pct_low', 'profit_per_grid_pct_high',
        'liq_price_long', 'liq_price_short', 'trigger_price'
    }
    filtered_data = {k: v for k, v in data.items() if k in valid_fields}

    bot_data = ExtractedBotData(**filtered_data)
    bot_data.compute_derived_fields()
    return bot_data


# =============================================================================
# CSV DATA EXTRACTION
# =============================================================================

def load_trade_history(csv_path: Path) -> pd.DataFrame:
    """Load trade history CSV."""
    if not csv_path.exists():
        logger.warning(f"Trade history CSV not found: {csv_path}")
        return pd.DataFrame()

    df = pd.read_csv(csv_path)
    logger.info(f"Loaded {len(df)} trades from {csv_path.name}")
    return df


def load_transaction_history(csv_path: Path) -> pd.DataFrame:
    """Load transaction history CSV (for funding fees)."""
    if not csv_path.exists():
        logger.warning(f"Transaction history CSV not found: {csv_path}")
        return pd.DataFrame()

    df = pd.read_csv(csv_path)
    logger.info(f"Loaded {len(df)} transactions from {csv_path.name}")
    return df


def filter_trades_by_symbol_and_time(
    df: pd.DataFrame,
    symbol: str,
    start_time: datetime,
    end_time: datetime
) -> pd.DataFrame:
    """Filter trades for a specific symbol and time range."""
    if df.empty:
        return df

    df = df.copy()
    time_col = 'Time(UTC)' if 'Time(UTC)' in df.columns else 'Time'
    df['parsed_time'] = pd.to_datetime(df[time_col], utc=True)

    mask = (
        (df['Symbol'] == symbol) &
        (df['parsed_time'] >= start_time) &
        (df['parsed_time'] <= end_time)
    )

    filtered = df[mask].copy()
    logger.info(f"Filtered to {len(filtered)} trades for {symbol} between {start_time} and {end_time}")
    return cast(pd.DataFrame, filtered)


def clean_numeric_column(series: pd.Series) -> pd.Series:
    """Clean numeric column values by removing currency suffixes like ' USDT'."""
    if series.dtype == 'object' or pd.api.types.is_string_dtype(series):
        cleaned = series.astype(str).str.replace(r'\s*USDT\s*$', '', regex=True)
        cleaned = cleaned.str.replace(r'\s*BTC\s*$', '', regex=True)
        cleaned = cleaned.str.replace(r',', '', regex=True)
        return cast(pd.Series, pd.to_numeric(cleaned, errors='coerce'))
    return cast(pd.Series, pd.to_numeric(series, errors='coerce'))


# =============================================================================
# TRADE METRICS
# =============================================================================

def _to_excursion_trade_events(trades_df: pd.DataFrame) -> List[Dict[str, Any]]:
    """Convert trade-history rows into MTM excursion engine trade events."""
    if trades_df.empty:
        return []

    if "parsed_time" in trades_df.columns:
        times = pd.to_datetime(trades_df["parsed_time"], utc=True, errors="coerce")
    else:
        time_col = "Time(UTC)" if "Time(UTC)" in trades_df.columns else "Time"
        if time_col not in trades_df.columns:
            return []
        times = pd.to_datetime(trades_df[time_col], utc=True, errors="coerce")

    side_col = "Side" if "Side" in trades_df.columns else "side"
    if side_col not in trades_df.columns:
        return []
    qty_col = "Quantity" if "Quantity" in trades_df.columns else "qty"
    price_col = "Price" if "Price" in trades_df.columns else "Avg. Price"
    if price_col not in trades_df.columns:
        return []
    realized_col = "Realized Profit" if "Realized Profit" in trades_df.columns else None
    fee_col = "Fee" if "Fee" in trades_df.columns else None
    pos_col = "Position Side" if "Position Side" in trades_df.columns else "positionSide"

    qty = clean_numeric_column(cast(pd.Series, trades_df[qty_col])) if qty_col in trades_df.columns else pd.Series(0.0, index=trades_df.index)
    prices = clean_numeric_column(cast(pd.Series, trades_df[price_col]))
    realized = clean_numeric_column(cast(pd.Series, trades_df[realized_col])) if realized_col else pd.Series(0.0, index=trades_df.index)
    fees = clean_numeric_column(cast(pd.Series, trades_df[fee_col])).abs() if fee_col else pd.Series(0.0, index=trades_df.index)

    events: List[Dict[str, Any]] = []
    for idx in trades_df.index:
        ts = times.loc[idx]
        side = str(trades_df.at[idx, side_col]).upper()
        q = float(qty.loc[idx]) if pd.notna(qty.loc[idx]) else 0.0
        price = float(prices.loc[idx]) if pd.notna(prices.loc[idx]) else 0.0
        if pd.isna(ts) or side not in {"BUY", "SELL"} or q <= 0 or price <= 0:
            continue
        raw_pos = trades_df.at[idx, pos_col] if pos_col in trades_df.columns else "BOTH"
        position_side = str(raw_pos).upper()
        if position_side not in {"LONG", "SHORT", "BOTH"}:
            position_side = "BOTH"

        events.append(
            {
                "time": int(cast(pd.Timestamp, pd.Timestamp(ts)).timestamp() * 1000),
                "side": side,
                "positionSide": position_side,
                "qty": q,
                "price": price,
                "realizedPnl": float(realized.loc[idx]) if pd.notna(realized.loc[idx]) else 0.0,
                "commission": float(fees.loc[idx]) if pd.notna(fees.loc[idx]) else 0.0,
            }
        )
    return events


def compute_trade_metrics(
    trades_df: pd.DataFrame,
    *,
    mark_klines: Optional[List[List[Any]]] = None,
    funding_fee_events: Optional[List[Dict[str, Any]]] = None,
    investment_margin_usdt: Optional[float] = None,
) -> TradeMetrics:
    """Compute metrics from filtered trade data.

    Backward-compatible signature — do not change parameter names or order.
    """
    metrics = TradeMetrics()

    if trades_df.empty:
        return metrics

    # Executed quantity
    if 'Quantity' in trades_df.columns:
        metrics.executed_qty = clean_numeric_column(cast(pd.Series, trades_df['Quantity'])).sum()

    # Commission (fees)
    if 'Fee' in trades_df.columns:
        metrics.commission_usdt = clean_numeric_column(cast(pd.Series, trades_df['Fee'])).abs().sum()

    # Maker/Taker counts
    if 'Maker' in trades_df.columns:
        maker_col = trades_df['Maker']
        if maker_col.dtype == 'bool':
            metrics.maker_count = int(maker_col.sum())
        else:
            metrics.maker_count = int((maker_col.astype(str).str.upper() == 'TRUE').sum())
        metrics.taker_count = len(trades_df) - metrics.maker_count

    # Notional value
    if 'Amount' in trades_df.columns:
        metrics.notional_completed = clean_numeric_column(cast(pd.Series, trades_df['Amount'])).abs().sum()
    elif 'Price' in trades_df.columns and 'Quantity' in trades_df.columns:
        prices = clean_numeric_column(cast(pd.Series, trades_df['Price']))
        quantities = clean_numeric_column(cast(pd.Series, trades_df['Quantity']))
        metrics.notional_completed = (prices * quantities).abs().sum()

    # Profit factor and win rate from Realized Profit
    if 'Realized Profit' in trades_df.columns:
        profits = clean_numeric_column(cast(pd.Series, trades_df['Realized Profit']))
        gross_profits = profits[profits > 0].sum()
        gross_losses = abs(profits[profits < 0].sum())

        metrics.gross_pnl = profits.sum()

        if gross_losses > 0:
            metrics.profit_factor = min(gross_profits / gross_losses, 999.99)
        elif gross_profits > 0:
            metrics.profit_factor = 999.99
        else:
            metrics.profit_factor = 0.0

        winning_trades = len(profits[profits > 0])
        total_trades = len(profits[profits != 0])
        if total_trades > 0:
            metrics.win_rate = (winning_trades / total_trades) * 100

    excursion = calculate_mtm_excursions(
        trades=_to_excursion_trade_events(trades_df),
        funding_fees=funding_fee_events or [],
        mark_klines=mark_klines or [],
    )
    metrics.mae = float(excursion.mae_usdt)
    metrics.mfe = float(excursion.mfe_usdt)
    metrics.mtm_used_mark_prices = excursion.used_mark_prices
    if investment_margin_usdt and investment_margin_usdt > 0:
        metrics.mae_pct_initial = (metrics.mae / investment_margin_usdt) * 100
        metrics.mfe_pct_initial = (metrics.mfe / investment_margin_usdt) * 100

    return metrics


def extract_funding_fee_events_from_transactions(
    tx_df: pd.DataFrame,
    symbol: str,
    start_time: datetime,
    end_time: datetime,
) -> List[Dict[str, Any]]:
    """Extract funding-fee cashflow events for MTM excursion accounting."""
    if tx_df.empty:
        return []

    time_col = 'Date(UTC)' if 'Date(UTC)' in tx_df.columns else 'Time'
    if time_col not in tx_df.columns:
        return []
    tx_df = tx_df.copy()
    tx_df['parsed_time'] = pd.to_datetime(tx_df[time_col], utc=True, errors='coerce')
    start_ts = pd.Timestamp(start_time)
    end_ts = pd.Timestamp(end_time)
    if start_ts.tzinfo is None:
        start_ts = start_ts.tz_localize("UTC")
    else:
        start_ts = start_ts.tz_convert("UTC")
    if end_ts.tzinfo is None:
        end_ts = end_ts.tz_localize("UTC")
    else:
        end_ts = end_ts.tz_convert("UTC")

    type_col = 'type' if 'type' in tx_df.columns else 'Type'
    if type_col not in tx_df.columns:
        return []

    amount_col = 'Amount' if 'Amount' in tx_df.columns else 'income'
    if amount_col not in tx_df.columns:
        return []

    type_mask = tx_df[type_col].astype(str).str.upper() == 'FUNDING_FEE'
    time_mask = (tx_df['parsed_time'] >= start_ts) & (tx_df['parsed_time'] <= end_ts)
    if 'Symbol' in tx_df.columns:
        symbol_mask = tx_df['Symbol'].astype(str).str.upper() == symbol.upper()
    elif 'symbol' in tx_df.columns:
        symbol_mask = tx_df['symbol'].astype(str).str.upper() == symbol.upper()
    else:
        symbol_mask = pd.Series(True, index=tx_df.index)

    filtered = tx_df[type_mask & time_mask & symbol_mask].copy()
    if filtered.empty:
        return []

    amounts = clean_numeric_column(cast(pd.Series, filtered[amount_col])).fillna(0.0)
    events: List[Dict[str, Any]] = []
    for ts, income in zip(filtered['parsed_time'], amounts):
        if pd.isna(ts):
            continue
        events.append({
            'time': int(cast(pd.Timestamp, pd.Timestamp(ts)).timestamp() * 1000),
            'income': float(income),
        })
    events.sort(key=lambda row: int(row['time']))
    return events


# =============================================================================
# BINANCE API DATA FETCHING
# =============================================================================

async def fetch_historical_klines(
    client: BinanceClient,
    symbol: str,
    start_time: datetime
) -> Dict[str, pd.DataFrame]:
    """Fetch historical klines for all timeframes at the bot's start time."""
    klines_data = {}
    start_ms = int(start_time.timestamp() * 1000)

    _kl = get_config().validation.kline_limits
    timeframes = {
        "1h": _kl.get("1h", 100),
        "15m": _kl.get("15m", 200),
        "5m": _kl.get("5m", 300),
        "1m": _kl.get("1m", 500),
    }

    for tf, limit in timeframes.items():
        try:
            klines = await client._request(
                get_config().binance.endpoints["klines"],
                params={
                    "symbol": symbol.upper(),
                    "interval": tf,
                    "endTime": start_ms,
                    "limit": limit,
                }
            )

            df = pd.DataFrame(klines, columns=cast(Any, [
                'open_time', 'open', 'high', 'low', 'close', 'volume',
                'close_time', 'quote_volume', 'trades', 'taker_buy_base',
                'taker_buy_quote', 'ignore'
            ]))

            for col in ['open', 'high', 'low', 'close', 'volume', 'quote_volume']:
                df[col] = pd.to_numeric(df[col], errors='coerce')

            klines_data[tf] = df
            logger.info(f"Fetched {len(df)} {tf} klines for {symbol}")

        except Exception as e:
            logger.error(f"Failed to fetch {tf} klines for {symbol}: {e}")
            klines_data[tf] = pd.DataFrame()

    return klines_data


async def fetch_session_mark_klines(
    client: BinanceClient,
    symbol: str,
    start_time: datetime,
    end_time: datetime,
) -> List[List[Any]]:
    """Fetch 1-minute mark-price klines for the full bot session."""
    start_ms = int(start_time.timestamp() * 1000)
    end_ms = int(end_time.timestamp() * 1000)
    if end_ms <= start_ms:
        return []
    try:
        klines = await client.get_mark_price_klines_all(
            symbol=symbol,
            interval="1m",
            start_time=start_ms,
            end_time=end_ms,
            include_current=False,
        )
        logger.info(f"Fetched {len(klines)} session mark-price klines for {symbol}")
        return klines
    except Exception as e:
        logger.error(f"Failed to fetch session mark-price klines for {symbol}: {e}")
        return []


# =============================================================================
# TECHNICAL INDICATOR COMPUTATION
# =============================================================================

def compute_indicators(klines: Dict[str, pd.DataFrame], bot_data: ExtractedBotData) -> TechnicalIndicators:
    """Compute technical indicators from historical klines."""
    _ind = get_config().indicators
    indicators = TechnicalIndicators()

    # 1H indicators
    df_1h = klines.get("1h", pd.DataFrame())
    if not df_1h.empty and len(df_1h) >= 50:
        closes = np.asarray(df_1h['close'])
        highs = np.asarray(df_1h['high'])
        lows = np.asarray(df_1h['low'])

        adx, _, _ = calc_adx(highs, lows, closes, _ind.adx_period)
        indicators.adx_1h = float(adx[-1]) if len(adx) > 0 and not np.isnan(adx[-1]) else None

        ema50 = calc_ema(closes, _ind.ema_medium)
        ema_slope = calc_ema_slope(ema50)
        indicators.ema_slope_1h = float(ema_slope[-1]) if len(ema_slope) > 0 and np.isfinite(ema_slope[-1]) else None

    # 15M indicators
    df_15m = klines.get("15m", pd.DataFrame())
    if not df_15m.empty and len(df_15m) >= 50:
        closes = np.asarray(df_15m['close'])
        highs = np.asarray(df_15m['high'])
        lows = np.asarray(df_15m['low'])
        volumes = np.asarray(df_15m['volume'])

        adx, _, _ = calc_adx(highs, lows, closes, _ind.adx_period)
        indicators.adx_15m = float(adx[-1]) if len(adx) > 0 and not np.isnan(adx[-1]) else None

        rsi = calc_rsi(closes, _ind.rsi_period)
        indicators.rsi_15m = float(rsi[-1]) if len(rsi) > 0 and not np.isnan(rsi[-1]) else None

        _, _, _, bb_width = calc_bollinger_bands(closes, _ind.bb_period, _ind.bb_std_dev)
        indicators.bb_width = float(bb_width[-1]) if len(bb_width) > 0 and np.isfinite(bb_width[-1]) else None

        atr = calc_atr(highs, lows, closes, _ind.atr_period)
        indicators.atr_15m = float(atr[-1]) if len(atr) > 0 and not np.isnan(atr[-1]) else None
        indicators.last_price = float(closes[-1])
        if indicators.atr_15m is not None and indicators.last_price is not None and indicators.last_price > 0:
            indicators.atr_pct_15m = (indicators.atr_15m / indicators.last_price) * 100

        indicators.trend_structure = detect_trend_structure(highs, lows)

    # 5M indicators
    df_5m = klines.get("5m", pd.DataFrame())
    if not df_5m.empty and len(df_5m) >= 50:
        closes = np.asarray(df_5m['close'])
        highs = np.asarray(df_5m['high'])
        lows = np.asarray(df_5m['low'])
        volumes = np.asarray(df_5m['volume'])

        adx, _, _ = calc_adx(highs, lows, closes, _ind.adx_period)
        indicators.adx_5m = float(adx[-1]) if len(adx) > 0 and not np.isnan(adx[-1]) else None

        ema20 = calc_ema(closes, _ind.ema_fast)
        ema50 = calc_ema(closes, _ind.ema_medium)
        indicators.ema_crosses_5m = count_ema_crosses(ema20[-100:], ema50[-100:])

        vwap = calc_vwap(highs, lows, closes, volumes)
        indicators.vwap_crosses_5m = count_vwap_crosses(closes[-100:], vwap[-100:])

    # Range size from Bollinger Bands
    if not df_15m.empty and len(df_15m) >= 50:
        upper_bb, _, lower_bb, _ = calc_bollinger_bands(np.asarray(df_15m['close']), _ind.bb_period, _ind.bb_std_dev)
        if len(upper_bb) > 0 and len(lower_bb) > 0:
            range_size = float(upper_bb[-1] - lower_bb[-1])
            last_close = float(df_15m['close'].values[-1])
            if last_close > 0 and np.isfinite(range_size):
                indicators.range_size_pct = (range_size / last_close) * 100

    return indicators


# =============================================================================
# CANDIDATE MATCHING
# =============================================================================

def search_candidate_match(
    symbol: str,
    grid_lower: float,
    grid_upper: float,
    grids_count: int,
    results_dir: Optional[Path] = None,
    tolerance_pct: float = 1.0,
) -> Optional[str]:
    """Search results/deployment_ready_*.csv files for a matching candidate_id.

    Match criteria (all must hold):
      - symbol exact match
      - grid_lower within tolerance_pct of price_range_low
      - grid_upper within tolerance_pct of price_range_high
      - num_grids == grids_count

    Returns the candidate_id string or None.
    """
    if results_dir is None:
        results_dir = Path("results")
    if not results_dir.exists():
        logger.warning("Results directory not found: %s", results_dir)
        return None

    csv_files = sorted(results_dir.glob("deployment_ready_*.csv"), reverse=True)
    if not csv_files:
        logger.info("No deployment_ready CSVs found in %s", results_dir)
        return None

    best_match: Optional[str] = None
    best_score = float("inf")

    for csv_path in csv_files:
        try:
            df = pd.read_csv(csv_path)
        except Exception:
            continue

        # Flexible column names
        sym_col = "symbol" if "symbol" in df.columns else None
        lower_col = next((c for c in df.columns if c in ("grid_lower", "price_range_low")), None)
        upper_col = next((c for c in df.columns if c in ("grid_upper", "price_range_high")), None)
        grids_col = next((c for c in df.columns if c in ("num_grids", "grids_count")), None)
        cid_col = "candidate_id" if "candidate_id" in df.columns else None

        if not all([sym_col, lower_col, upper_col, grids_col, cid_col]):
            continue

        # Filter by symbol
        sym_mask = df[sym_col].astype(str).str.upper() == symbol.upper()
        candidates = df[sym_mask]
        if candidates.empty:
            continue

        for _, row in candidates.iterrows():
            try:
                r_lower = float(row[lower_col])
                r_upper = float(row[upper_col])
                r_grids = int(row[grids_col])
            except (ValueError, TypeError):
                continue

            if r_grids != grids_count:
                continue

            # Check price tolerance
            if grid_lower > 0:
                lower_diff_pct = abs(r_lower - grid_lower) / grid_lower * 100
            else:
                lower_diff_pct = float("inf")
            if grid_upper > 0:
                upper_diff_pct = abs(r_upper - grid_upper) / grid_upper * 100
            else:
                upper_diff_pct = float("inf")

            if lower_diff_pct > tolerance_pct or upper_diff_pct > tolerance_pct:
                continue

            score = lower_diff_pct + upper_diff_pct
            if score < best_score:
                best_score = score
                cid = cast(Any, row[cid_col])
                if pd.notna(cid) and str(cid).strip():
                    best_match = str(cid).strip()

    if best_match:
        logger.info("Candidate match found: %s (score=%.4f)", best_match, best_score)
    else:
        logger.info("No candidate match found for %s [%.2f-%.2f, %d grids]",
                     symbol, grid_lower, grid_upper, grids_count)
    return best_match


# =============================================================================
# DATA COHERENCE VALIDATION
# =============================================================================

def validate_data_coherence(
    bot: ExtractedBotData,
    trade_metrics: TradeMetrics,
    fills: Optional[List[ParsedTradeFill]] = None,
) -> List[str]:
    """Cross-validate data consistency. Returns list of warning strings."""
    warnings: List[str] = []

    # 1. PnL % vs total_profit / margin
    if (bot.pnl_pct is not None
            and bot.total_profit_usdt is not None
            and bot.invested_margin_usdt
            and bot.invested_margin_usdt > 0):
        expected_pct = (bot.total_profit_usdt / bot.invested_margin_usdt) * 100
        diff = abs(expected_pct - bot.pnl_pct)
        if diff > 1.0:
            warnings.append(
                f"PnL% mismatch: reported={bot.pnl_pct:.2f}%, "
                f"computed={expected_pct:.2f}% (diff={diff:.2f}%)"
            )

    # 2. Total profit ≈ matched + unmatched + funding
    if (bot.total_profit_usdt is not None
            and bot.realized_pnl_usdt is not None
            and bot.unrealized_pnl_usdt is not None):
        funding = bot.funding_fee_usdt or 0.0
        expected_total = bot.realized_pnl_usdt + bot.unrealized_pnl_usdt + funding
        diff = abs(expected_total - bot.total_profit_usdt)
        if diff > 0.5:
            warnings.append(
                f"Profit decomposition mismatch: total={bot.total_profit_usdt:.4f}, "
                f"matched+unmatched+funding={expected_total:.4f} (diff={diff:.4f})"
            )

    # 3. Commission from fills vs trade_metrics
    if fills and trade_metrics.commission_usdt > 0:
        fill_fees = sum(f.fee for f in fills)
        diff = abs(fill_fees - trade_metrics.commission_usdt)
        if diff > 0.01:
            warnings.append(
                f"Fee mismatch: fills_sum={fill_fees:.6f}, "
                f"metrics={trade_metrics.commission_usdt:.6f}"
            )

    # 4. Grid spacing sanity
    if bot.grid_spacing_pct is not None:
        if bot.grid_spacing_pct <= 0:
            warnings.append(f"Grid spacing non-positive: {bot.grid_spacing_pct:.4f}%")
        elif bot.grid_spacing_pct > 20:
            warnings.append(f"Grid spacing unusually large: {bot.grid_spacing_pct:.4f}%")

    # 5. Liquidation sanity: grid should be inside liq bounds
    if bot.liq_price_long is not None and bot.price_range_low is not None:
        if bot.liq_price_long >= bot.price_range_low:
            warnings.append(
                f"Liq long ({bot.liq_price_long}) >= grid low ({bot.price_range_low})"
            )
    if bot.liq_price_short is not None and bot.price_range_high is not None:
        if bot.liq_price_short <= bot.price_range_high:
            warnings.append(
                f"Liq short ({bot.liq_price_short}) <= grid high ({bot.price_range_high})"
            )

    for w in warnings:
        logger.warning("COHERENCE: %s", w)

    return warnings


def auto_repair_coherence(
    bot: ExtractedBotData,
    trade_metrics: TradeMetrics,
    warnings: List[str],
) -> tuple[List[str], List[str]]:
    """Attempt automatic repair of coherence issues.

    Returns (repaired_list, unresolved_list) — both lists of human-readable strings.
    Mutates ``bot`` and ``trade_metrics`` in place for repairable issues.
    """
    repaired: List[str] = []
    unresolved: List[str] = []

    for w in warnings:
        # ── holding_time_ok recomputation ────────────────────────────
        # Not a direct coherence warning but always re-derive from duration_hours
        # (handled below, outside the per-warning loop)

        # ── Profit decomposition mismatch ────────────────────────────
        if "Profit decomposition mismatch" in w:
            # If all three components are present, recalculate total_profit_usdt
            if (bot.realized_pnl_usdt is not None
                    and bot.unrealized_pnl_usdt is not None):
                funding = bot.funding_fee_usdt or 0.0
                recomputed = bot.realized_pnl_usdt + bot.unrealized_pnl_usdt + funding
                # Only auto-repair if the Binance-reported components are credible
                # (i.e. total is further from zero than the sum — likely a display rounding)
                if bot.total_profit_usdt is not None:
                    diff = abs(recomputed - bot.total_profit_usdt)
                    if diff <= 5.0:
                        # Minor mismatch — Binance total is authoritative, no data changed
                        repaired.append(
                            f"Profit decomposition diff {diff:.4f} USDT within 5.0 tolerance "
                            f"(no correction needed)"
                        )
                    else:
                        # Large mismatch — flag as unresolved, don't silently overwrite
                        unresolved.append(w)
                else:
                    unresolved.append(w)
            else:
                unresolved.append(w)

        # ── PnL % mismatch ───────────────────────────────────────────
        elif "PnL% mismatch" in w:
            # PnL% = total_profit / margin * 100.  Recalculate from source.
            if (bot.total_profit_usdt is not None
                    and bot.invested_margin_usdt
                    and bot.invested_margin_usdt > 0):
                recomputed_pct = (bot.total_profit_usdt / bot.invested_margin_usdt) * 100
                bot.pnl_pct = round(recomputed_pct, 2)
                repaired.append(
                    f"PnL% recomputed from total_profit/margin: {bot.pnl_pct:.2f}%"
                )
            else:
                unresolved.append(w)

        # ── Grid spacing non-positive / unusually large ──────────────
        elif "Grid spacing" in w:
            # Cannot auto-repair grid spacing — user-entered data
            unresolved.append(w)

        # ── Liquidation sanity ───────────────────────────────────────
        elif "Liq long" in w or "Liq short" in w:
            # Liquidation prices crossing grid bounds usually means
            # the OCR/parsing extracted them incorrectly.  Flag only.
            unresolved.append(w)

        # ── Fee mismatch ─────────────────────────────────────────────
        elif "Fee mismatch" in w:
            # Minor fee rounding — tolerate
            unresolved.append(w)

        else:
            unresolved.append(w)

    # holding_time_ok is always re-derived from duration_hours in
    # prepare_entry_validation_row(), so no repair action needed here.

    return repaired, unresolved


# =============================================================================
# LIQUIDATION FEATURES
# =============================================================================

def _compute_liquidation_features(
    liq_long: Optional[float],
    liq_short: Optional[float],
    grid_low: Optional[float],
    grid_high: Optional[float],
) -> Dict[str, Any]:
    """Compute liquidation distance features from liq prices and grid bounds.

    Distances use grid bounds (NOT trigger_price) as reference.
    Asymmetry is dist_short / dist_long (ratio, NOT normalized difference).
    """
    features: Dict[str, Any] = {
        "liq_price_long": liq_long,
        "liq_price_short": liq_short,
        "dist_to_liq_long_pct": None,
        "dist_to_liq_short_pct": None,
        "liq_range_utilization": None,
        "liq_asymmetry": None,
    }

    if liq_long is not None and grid_low is not None and grid_low > 0:
        features["dist_to_liq_long_pct"] = round(
            (grid_low - liq_long) / grid_low * 100, 4
        )

    if liq_short is not None and grid_high is not None and grid_high > 0:
        features["dist_to_liq_short_pct"] = round(
            (liq_short - grid_high) / grid_high * 100, 4
        )

    if (
        liq_long is not None
        and liq_short is not None
        and grid_low is not None
        and grid_high is not None
        and liq_short > liq_long
    ):
        liq_range = liq_short - liq_long
        grid_range = grid_high - grid_low
        features["liq_range_utilization"] = round(
            (grid_range / liq_range) * 100, 4
        )

    dist_long = features.get("dist_to_liq_long_pct")
    dist_short = features.get("dist_to_liq_short_pct")
    if dist_long is not None and dist_short is not None and dist_long > 0:
        features["liq_asymmetry"] = round(dist_short / dist_long, 4)

    return features


# =============================================================================
# ROW PREPARATION & EXCEL WRITER
# =============================================================================

_START_TIME_CANONICAL_FORMAT = "%Y-%m-%d %H:%M:%S"


def normalize_start_time_utc(value: Any) -> Optional[str]:
    """Canonical `start_time_utc` for workbook writes.

    Returns a fixed-width `YYYY-MM-DD HH:MM:SS` UTC string so that lexicographic
    sort equals chronological sort and the column is immune to Excel
    datetime reinterpretation on save/open cycles. Accepts Python datetime
    (tz-aware or naive, treated as UTC), pandas Timestamp, or an already-formatted
    string. Returns None for None.
    """
    if value is None:
        return None
    if isinstance(value, str):
        if not value.strip():
            return None
        ts = pd.to_datetime(value, errors="coerce", utc=True)
        if pd.isna(ts):
            raise ValueError(f"Unparseable start_time_utc string: {value!r}")
        return ts.strftime(_START_TIME_CANONICAL_FORMAT)
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value.strftime(_START_TIME_CANONICAL_FORMAT)
    ts = pd.to_datetime(value, errors="coerce", utc=True)
    if pd.isna(ts):
        raise ValueError(f"Unsupported start_time_utc type: {type(value).__name__}")
    return ts.strftime(_START_TIME_CANONICAL_FORMAT)


def prepare_entry_validation_row(
    bot: ExtractedBotData,
    indicators: TechnicalIndicators,
    trade_metrics: TradeMetrics,
    candidate_id: Optional[str] = None,
) -> Dict[str, Any]:
    """Prepare a row for the Entry Validation Metrics sheet."""
    liq_features = _compute_liquidation_features(
        bot.liq_price_long, bot.liq_price_short,
        bot.price_range_low, bot.price_range_high,
    )
    return {
        "strategy_id": bot.strategy_id,
        "symbol": bot.symbol,
        "strategy_type": bot.strategy_type,
        "status": bot.status,
        "start_time_utc": normalize_start_time_utc(bot.start_time_utc),
        "end_time_utc": bot.end_time_utc,
        "duration_hours": round(bot.duration_hours, 2) if bot.duration_hours is not None else None,
        "invested_margin_usdt": bot.invested_margin_usdt,
        "leverage": bot.leverage,
        "grids_count": bot.grids_count,
        "mode": bot.mode,
        "price_range_low": bot.price_range_low,
        "price_range_high": bot.price_range_high,
        "grid_spacing_pct": round(bot.grid_spacing_pct, 4) if bot.grid_spacing_pct is not None else None,
        "total_profit_usdt": bot.total_profit_usdt,
        "pnl_pct": bot.pnl_pct,
        "realized_pnl_usdt": bot.realized_pnl_usdt,
        "unrealized_pnl_usdt": bot.unrealized_pnl_usdt,
        "funding_fee_usdt": bot.funding_fee_usdt,
        "commission_usdt": trade_metrics.commission_usdt,
        "mae": trade_metrics.mae,
        "mfe": trade_metrics.mfe,
        "mae_pct_initial": trade_metrics.mae_pct_initial,
        "mfe_pct_initial": trade_metrics.mfe_pct_initial,
        "profit_factor": trade_metrics.profit_factor,
        "maker_count": trade_metrics.maker_count,
        "taker_count": trade_metrics.taker_count,
        "total_trades": bot.total_trades or (trade_metrics.maker_count + trade_metrics.taker_count),
        "executed_qty": trade_metrics.executed_qty,
        "notional_completed": trade_metrics.notional_completed,
        "holding_time_ok": int((bot.duration_hours or 0) <= 48),
        "last_updated_utc": datetime.now(timezone.utc),
        "adx_1h": round(indicators.adx_1h, 2) if indicators.adx_1h is not None else None,
        "adx_15m": round(indicators.adx_15m, 2) if indicators.adx_15m is not None else None,
        "adx_5m": round(indicators.adx_5m, 2) if indicators.adx_5m is not None else None,
        "rsi_15m": round(indicators.rsi_15m, 2) if indicators.rsi_15m is not None else None,
        "ema_slope_1h": indicators.ema_slope_1h,
        "ema_crosses_5m": indicators.ema_crosses_5m,
        "vwap_crosses_5m": indicators.vwap_crosses_5m,
        "range_size_pct": round(indicators.range_size_pct, 4) if indicators.range_size_pct is not None else None,
        "bb_width": round(indicators.bb_width, 4) if indicators.bb_width is not None else None,
        "trend_structure": indicators.trend_structure,
        "candidate_id": candidate_id,
        **liq_features,
        "trigger_price": bot.trigger_price,
    }


def append_to_excel(output_path: Path, row: Dict[str, Any]):
    """Append a single row to the canonical workbook (Sheet1)."""
    import openpyxl

    CANONICAL_COLUMNS = [
        "strategy_id", "symbol", "strategy_type", "status",
        "start_time_utc", "end_time_utc", "duration_hours",
        "invested_margin_usdt", "leverage", "grids_count",
        "price_range_low", "price_range_high", "grid_spacing_pct",
        "total_profit_usdt", "pnl_pct", "realized_pnl_usdt",
        "unrealized_pnl_usdt", "funding_fee_usdt", "commission_usdt",
        "profit_factor", "maker_count", "taker_count", "total_trades",
        "executed_qty", "notional_completed", "holding_time_ok",
        "last_updated_utc", "adx_1h", "adx_15m", "adx_5m",
        "rsi_15m", "ema_slope_1h", "ema_crosses_5m", "vwap_crosses_5m",
        "range_size_pct", "bb_width", "trend_structure", "candidate_id",
        "mae", "mfe", "mae_pct_initial", "mfe_pct_initial",
        "liq_price_long", "liq_price_short",
        "dist_to_liq_long_pct", "dist_to_liq_short_pct",
        "liq_range_utilization", "liq_asymmetry",
        "trigger_price",
        "coherence_ok", "coherence_warnings",
    ]

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        assert ws is not None
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        headers = [cell.value for cell in header_row]
        # ERR-069: fail closed instead of widening a foreign schema in place.
        # This legacy writer may only append to workbooks it created itself
        # (header exactly CANONICAL_COLUMNS). The canonical training workbook
        # is managed by new_bot_data_extractor.py (resolve_workbook_columns /
        # append_row_preserving_schema) and must never be widened here.
        if headers != CANONICAL_COLUMNS:
            wb.close()
            raise ValueError(
                "append_to_excel: existing workbook header does not match the "
                f"legacy CANONICAL_COLUMNS schema ({output_path}). Use "
                "new_bot_data_extractor.py for the canonical workbook."
            )
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        headers = list(CANONICAL_COLUMNS)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

    # Duplicate check by strategy_id
    strategy_id = row.get("strategy_id")
    if strategy_id is not None:
        for r in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if r[0] == strategy_id:
                logger.warning("strategy_id %s already exists — skipping append", strategy_id)
                wb.close()
                return

    new_row_values = []
    for header in headers:
        val = row.get(cast(str, header))
        if header == "start_time_utc":
            val = normalize_start_time_utc(val)
        elif isinstance(val, datetime) and val.tzinfo is not None:
            val = val.replace(tzinfo=None)
        if isinstance(val, (bool, np.bool_)):
            val = int(val)
        new_row_values.append(val)

    ws.append(new_row_values)
    wb.save(output_path)
    wb.close()
    logger.info("Appended row to %s", output_path)


# =============================================================================
# MAIN PROCESSING
# =============================================================================

async def process_bot(
    bot_data: ExtractedBotData,
    trade_csv: Optional[Path] = None,
    txn_csv: Optional[Path] = None,
    output_path: Path = Path("data/new_expired_bots.xlsx"),
    verify: bool = True,
    trade_fills: Optional[List[ParsedTradeFill]] = None,
    pnl_curve: Optional[List[float]] = None,
    results_dir: Optional[Path] = None,
):
    """Process a single bot and write to Excel.

    Parameters
    ----------
    bot_data : ExtractedBotData
        Parsed bot parameters.
    trade_csv : Path, optional
        Path to Binance trade history CSV.
    txn_csv : Path, optional
        Path to transaction history CSV (funding fees).
    output_path : Path
        Output Excel file.
    verify : bool
        Display extracted data for user verification.
    trade_fills : list of ParsedTradeFill, optional
        Trade fills parsed from user text (used when no trade_csv).
    pnl_curve : list of float, optional
        Cumulative PnL curve (fallback for MAE/MFE when no mark klines).
    results_dir : Path, optional
        Directory to search for candidate_id matching.
    """
    if not bot_data.symbol:
        raise ValueError("Symbol is required")

    if not bot_data.start_time_utc or not bot_data.end_time_utc:
        raise ValueError("Start and end times are required")

    if verify:
        bot_data.display_verification()

    logger.info(f"Processing bot: {bot_data.symbol} (Strategy ID: {bot_data.strategy_id})")
    logger.info(f"  Time range: {bot_data.start_time_utc} to {bot_data.end_time_utc}")

    # --- Candidate matching ---
    candidate_id: Optional[str] = None
    if (bot_data.price_range_low and bot_data.price_range_high and bot_data.grids_count):
        candidate_id = search_candidate_match(
            symbol=bot_data.symbol,
            grid_lower=bot_data.price_range_low,
            grid_upper=bot_data.price_range_high,
            grids_count=bot_data.grids_count,
            results_dir=results_dir,
        )

    # --- Binance API ---
    client = BinanceClient()

    try:
        logger.info("Fetching historical klines...")
        klines = await fetch_historical_klines(client, bot_data.symbol, bot_data.start_time_utc)

        logger.info("Fetching session mark-price klines...")
        session_mark_klines = await fetch_session_mark_klines(
            client, bot_data.symbol, bot_data.start_time_utc, bot_data.end_time_utc
        )

        # --- Technical indicators ---
        logger.info("Computing technical indicators...")
        indicators = compute_indicators(klines, bot_data)

        # --- Funding fees ---
        funding_fee_events: List[Dict[str, Any]] = []
        if txn_csv and txn_csv.exists():
            logger.info("Processing transaction history...")
            txn_df = load_transaction_history(txn_csv)
            funding_fee_events = extract_funding_fee_events_from_transactions(
                txn_df, bot_data.symbol, bot_data.start_time_utc, bot_data.end_time_utc
            )
            funding_from_txn = float(sum(float(event.get("income", 0.0)) for event in funding_fee_events))
            if funding_from_txn != 0.0 and bot_data.funding_fee_usdt is None:
                bot_data.funding_fee_usdt = funding_from_txn

        # --- Trade metrics ---
        trade_metrics = TradeMetrics()

        if trade_csv and trade_csv.exists():
            # Primary: CSV trade history
            logger.info("Processing trade history CSV...")
            trades_df = load_trade_history(trade_csv)
            filtered_trades = filter_trades_by_symbol_and_time(
                trades_df, bot_data.symbol, bot_data.start_time_utc, bot_data.end_time_utc
            )
            trade_metrics = compute_trade_metrics(
                filtered_trades,
                mark_klines=session_mark_klines,
                funding_fee_events=funding_fee_events,
                investment_margin_usdt=bot_data.invested_margin_usdt,
            )
        elif trade_fills:
            # Secondary: text-parsed fills
            logger.info("Computing metrics from %d parsed trade fills...", len(trade_fills))
            fills_df = text_fills_to_dataframe(trade_fills, bot_data.start_time_utc)
            trade_metrics = compute_trade_metrics(
                fills_df,
                mark_klines=session_mark_klines,
                funding_fee_events=funding_fee_events,
                investment_margin_usdt=bot_data.invested_margin_usdt,
            )

            # For text fills: compute profit_factor from fills if not set
            if trade_metrics.profit_factor is None and trade_fills:
                wins = sum(f.realized_profit for f in trade_fills if f.realized_profit > 0)
                losses = abs(sum(f.realized_profit for f in trade_fills if f.realized_profit < 0))
                if losses > 0:
                    trade_metrics.profit_factor = min(wins / losses, 999.99)
                elif wins > 0:
                    trade_metrics.profit_factor = 999.99
                else:
                    trade_metrics.profit_factor = 0.0

        # --- PnL curve fallback for MAE/MFE ---
        # Prefer PnL-curve excursions when:
        #   (a) MTM engine had no mark prices (used_mark_prices=False), OR
        #   (b) MTM MAE/MFE is None or 0.0 (degenerate result)
        # This avoids silently keeping zero excursions when a user-supplied
        # PnL curve provides richer information.
        if pnl_curve and len(pnl_curve) >= 2:
            curve_mae, curve_mfe = compute_pnl_curve_excursions(pnl_curve)
            mtm_unreliable = not trade_metrics.mtm_used_mark_prices
            if (mtm_unreliable or trade_metrics.mae is None or trade_metrics.mae == 0.0) and curve_mae > 0:
                trade_metrics.mae = curve_mae
                logger.info("MAE from PnL curve: %.4f (mtm_mark=%s)", curve_mae, trade_metrics.mtm_used_mark_prices)
            if (mtm_unreliable or trade_metrics.mfe is None or trade_metrics.mfe == 0.0) and curve_mfe > 0:
                trade_metrics.mfe = curve_mfe
                logger.info("MFE from PnL curve: %.4f (mtm_mark=%s)", curve_mfe, trade_metrics.mtm_used_mark_prices)
            # Recompute pct if margin available
            if bot_data.invested_margin_usdt and bot_data.invested_margin_usdt > 0:
                if trade_metrics.mae and trade_metrics.mae_pct_initial in (None, 0.0):
                    trade_metrics.mae_pct_initial = (trade_metrics.mae / bot_data.invested_margin_usdt) * 100
                if trade_metrics.mfe and trade_metrics.mfe_pct_initial in (None, 0.0):
                    trade_metrics.mfe_pct_initial = (trade_metrics.mfe / bot_data.invested_margin_usdt) * 100

        # --- Data coherence validation + auto-repair ---
        coherence_warnings = validate_data_coherence(
            bot_data, trade_metrics, trade_fills
        )
        repaired_msgs: List[str] = []
        unresolved_msgs: List[str] = []
        if coherence_warnings:
            repaired_msgs, unresolved_msgs = auto_repair_coherence(
                bot_data, trade_metrics, coherence_warnings
            )
            for msg in repaired_msgs:
                logger.info("COHERENCE REPAIRED: %s", msg)
            for msg in unresolved_msgs:
                logger.warning("COHERENCE UNRESOLVED: %s", msg)

        # --- Build row ---
        row = prepare_entry_validation_row(bot_data, indicators, trade_metrics, candidate_id)
        # Stamp coherence status onto the row
        row["coherence_ok"] = int(len(unresolved_msgs) == 0)
        row["coherence_warnings"] = "; ".join(unresolved_msgs) if unresolved_msgs else None

        # --- Write ---
        if unresolved_msgs:
            logger.warning(
                "Writing row with %d unresolved coherence issue(s) — "
                "coherence_ok=0 flagged in output.",
                len(unresolved_msgs),
            )
        logger.info("Writing to Excel...")
        append_to_excel(output_path, row)

        logger.info(f"Successfully processed {bot_data.symbol}")

        # Print summary
        print("\n" + "=" * 60)
        print(f"Bot Data Extraction Complete: {bot_data.symbol}")
        print("=" * 60)
        print(f"Strategy ID:     {bot_data.strategy_id}")
        print(f"Status:          {bot_data.status}")
        print(f"Duration:        {bot_data.duration_hours:.2f} hours" if bot_data.duration_hours else "Duration: N/A")
        print(f"Total Profit:    {bot_data.total_profit_usdt} USDT ({bot_data.pnl_pct}%)")
        print(f"ADX (1h/15m/5m): {indicators.adx_1h}/{indicators.adx_15m}/{indicators.adx_5m}")
        print(f"RSI (15m):       {indicators.rsi_15m}")
        print(f"Trend:           {indicators.trend_structure}")
        print(f"Trades:          {bot_data.total_trades}")
        print(f"Candidate ID:    {candidate_id or 'None'}")
        print(f"MAE/MFE:         {trade_metrics.mae}/{trade_metrics.mfe}")
        if coherence_warnings:
            print(f"Coherence:       {len(coherence_warnings)} issue(s)")
            if repaired_msgs:
                print(f"  Repaired:      {len(repaired_msgs)}")
            if unresolved_msgs:
                print(f"  Unresolved:    {len(unresolved_msgs)}")
                for w in unresolved_msgs:
                    print(f"  ! {w}")
        print(f"Output:          {output_path}")
        print("=" * 60)

    finally:
        await client.close()


# =============================================================================
# CLI
# =============================================================================

def parse_args() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Extract grid bot data and write to new_expired_bots.xlsx"
    )

    # Text input (PRIMARY)
    parser.add_argument("--text-file", type=Path, help="Path to text file with bot data")
    parser.add_argument("--text-input", action="store_true",
                        help="Read bot data from stdin (paste, then Ctrl+Z Enter / Ctrl+D)")

    # Image inputs (legacy)
    parser.add_argument("--image1", type=Path, help="Path to first screenshot")
    parser.add_argument("--image2", type=Path, help="Path to second screenshot")
    parser.add_argument("--image3", type=Path, help="Path to third screenshot")

    # Manual JSON input (legacy)
    parser.add_argument("--manual-input", type=Path, help="Path to JSON with manually extracted data")

    # CSV files
    parser.add_argument("--trade-csv", type=Path, help="Path to Trade History CSV")
    parser.add_argument("--txn-csv", type=Path, help="Path to Transaction History CSV")

    # Output
    parser.add_argument(
        "--output", type=Path,
        default=Path("data/new_expired_bots.xlsx"),
        help="Path to output Excel file"
    )

    # Results directory for candidate matching
    parser.add_argument(
        "--results-dir", type=Path, default=Path("results"),
        help="Directory with deployment_ready CSVs for candidate matching"
    )

    # Flags
    parser.add_argument("--no-verify", action="store_true",
                        help="Skip verification display")

    return parser.parse_args()


def main():
    """Main entry point."""
    args = parse_args()

    trade_fills: Optional[List[ParsedTradeFill]] = None
    pnl_curve: Optional[List[float]] = None

    # --- Input selection ---
    if args.text_file:
        if not args.text_file.exists():
            logger.error(f"Text file not found: {args.text_file}")
            sys.exit(1)
        text = args.text_file.read_text(encoding="utf-8")
        bot_data, trade_fills, pnl_curve = build_bot_data_from_text(text)
        logger.info(f"Loaded text input from {args.text_file}")

    elif args.text_input:
        print("Paste bot data below, then press Ctrl+Z Enter (Windows) or Ctrl+D (Unix):")
        text = sys.stdin.read()
        bot_data, trade_fills, pnl_curve = build_bot_data_from_text(text)
        logger.info("Loaded text input from stdin")

    elif args.manual_input:
        if not args.manual_input.exists():
            logger.error(f"Manual input file not found: {args.manual_input}")
            sys.exit(1)
        bot_data = load_manual_input(args.manual_input)
        logger.info(f"Loaded manual input from {args.manual_input}")

    elif args.image1 or args.image2 or args.image3:
        image_paths = [p for p in [args.image1, args.image2, args.image3] if p]
        bot_data = extract_from_images(image_paths)
        logger.info(f"Extracted data from {len(image_paths)} images")

    else:
        logger.error("Provide one of: --text-file, --text-input, --manual-input, or --image1/2/3")
        sys.exit(1)

    # --- Validate required fields ---
    if not bot_data.symbol:
        logger.error("Symbol could not be extracted. Check input data.")
        sys.exit(1)

    if not bot_data.end_time_utc or not bot_data.duration_hours:
        logger.error("End time and duration are required. Check input data.")
        sys.exit(1)

    # --- Run async processing ---
    asyncio.run(process_bot(
        bot_data,
        trade_csv=args.trade_csv,
        txn_csv=args.txn_csv,
        output_path=args.output,
        verify=not args.no_verify,
        trade_fills=trade_fills,
        pnl_curve=pnl_curve,
        results_dir=args.results_dir,
    ))


if __name__ == "__main__":
    main()
